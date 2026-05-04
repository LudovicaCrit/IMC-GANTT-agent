"""
IMC-Group GANTT Agent — Backend API (FastAPI)
Espone endpoint REST per il frontend React.
"""

from fastapi import FastAPI, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
from scenario_engine import simula_scenario, risultato_per_api, _to_date as scenario_to_date
from auth_routes import router as auth_router
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from deps import get_current_user, require_manager
from models import (
    Ruolo, Competenza, DipendentiCompetenze, FaseStandard, Fase, Utente, get_session, Dipendente,
)

import data as data_module
from data import (
    get_dipendente, get_tasks_progetto, get_progetti_dipendente,
    carico_settimanale_dipendente, ore_consuntivate_progetto,
    tasso_compilazione_progetto,
    aggiungi_task, modifica_task, cambia_stato_progetto,
    calcola_impatto_saturazione, get_progetto,
)

# Accesso ai DataFrame sempre via modulo (non copie statiche)
# Così dopo _reload() vediamo i dati aggiornati
def _DIPENDENTI(): return data_module.DIPENDENTI
def _PROGETTI(): return data_module.PROGETTI
def _TASKS(): return data_module.TASKS
def _CONSUNTIVI(): return data_module.CONSUNTIVI

# Prova a importare le funzioni persistenti (db mode)
try:
    from data import get_segnalazioni, aggiungi_segnalazione
    from data import salva_bozza_pianificazione, carica_bozza_pianificazione
    from data import salva_consuntivo
    PERSISTENT_MODE = True
except ImportError:
    PERSISTENT_MODE = False
from agent import (
    init_gemini, costruisci_contesto, chiedi_agente, is_agent_available,
)

app = FastAPI(title="IMC-Group GANTT Agent", version="0.1.0")

# Rate limiter - strategia: per IP del client
limiter = Limiter(key_func=get_remote_address)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

app.include_router(auth_router)

# CORS per permettere al frontend React di chiamare il backend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000", "http://localhost:5173"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_oggi():
    return datetime.now()

# ── Cache contesto per IA ──
_contesto_cache = {
    "data": None,
    "timestamp": None,
}

def get_contesto_ia():
    """Restituisce il contesto per l'IA, ricalcolandolo solo se i dati sono cambiati."""
    import json as json_mod
    
    # Invalida cache ogni 60 secondi (o dopo modifica dati)
    now = datetime.now()
    if (_contesto_cache["data"] is not None 
        and _contesto_cache["timestamp"] 
        and (now - _contesto_cache["timestamp"]).seconds < 60):
        return _contesto_cache["data"]
    
    # Ricostruisci contesto
    progetti_ctx = []
    for _, p in _PROGETTI().iterrows():
        if p["stato"] not in ("In esecuzione", "In bando"):
            continue
        tasks_prog = _TASKS()[_TASKS()["progetto_id"] == p["id"]]
        tasks_list = []
        for _, t in tasks_prog.iterrows():
            dip_nome = get_dipendente(t["dipendente_id"])["nome"] if t["dipendente_id"] else "Non assegnato"
            tasks_list.append({
                "id": t["id"], "nome": t["nome"],
                "assegnato_a": dip_nome,
                "inizio": t["data_inizio"].strftime("%Y-%m-%d") if t["data_inizio"] else "",
                "fine": t["data_fine"].strftime("%Y-%m-%d") if t["data_fine"] else "",
                "ore_stimate": int(t["ore_stimate"]),
                "stato": t["stato"],
            })
        progetti_ctx.append({
            "id": p["id"], "nome": p["nome"],
            "cliente": p["cliente"], "stato": p["stato"],
            "scadenza": p["data_fine"].strftime("%Y-%m-%d") if p["data_fine"] else "",
            "task": tasks_list,
        })

    dipendenti_ctx = []
    for _, d in _DIPENDENTI().iterrows():
        carico = carico_settimanale_dipendente(d["id"], get_oggi())
        dipendenti_ctx.append({
            "id": d["id"], "nome": d["nome"],
            "profilo": d["profilo"],
            "ore_sett": int(d["ore_sett"]),
            "saturazione_pct": round(carico / d["ore_sett"] * 100),
        })

    contesto = {
        "data_corrente": get_oggi().strftime("%Y-%m-%d"),
        "progetti": progetti_ctx,
        "dipendenti": dipendenti_ctx,
    }
    
    _contesto_cache["data"] = contesto
    _contesto_cache["timestamp"] = now
    
    return contesto

# ── Store segnalazioni in memoria (in futuro: database) ──
SEGNALAZIONI_STORE = []
_segn_counter = 0

def _next_segn_id():
    global _segn_counter
    _segn_counter += 1
    return f"S{_segn_counter:03d}"


# ══════════════════════════════════════════════════════════════════════
# MODELLI REQUEST/RESPONSE
# ══════════════════════════════════════════════════════════════════════
 
class AttivitaInternaRequest(BaseModel):
    dipendente_id: str
    nome: str
    categoria: str = "Formazione"
    ore_settimanali: int = 4
    ore_stimate: int = 0
    data_inizio: str
    data_fine: str
    note: str = ""


class ChatRequest(BaseModel):
    dipendente_id: str
    messaggio: str
    ore_compilate: dict[str, float] = {}
    stati_compilati: dict[str, str] = {}
    ore_assenza: float = 0.0
    tipo_assenza: str = ""
    nota_assenza: str = ""
    spese: list[dict] = []
    chat_history: list[dict] = []  # [{"role": "user"|"assistant", "content": "..."}]


class SimulaRitardoRequest(BaseModel):
    task_id: str
    giorni_ritardo: int


class RitardoItem(BaseModel):
    task_id: str
    giorni_ritardo: int


class SimulaRitardoMultiploRequest(BaseModel):
    ritardi: list[RitardoItem]


class SimulaRiassegnaRequest(BaseModel):
    task_id: str
    nuovo_dipendente_id: str


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: _DIPENDENTI()
# ══════════════════════════════════════════════════════════════════════

@app.get("/api/dipendenti")
def lista_dipendenti(_: Utente = Depends(require_manager)):
    result = []
    for _, d in _DIPENDENTI().iterrows():
        carico = carico_settimanale_dipendente(d["id"], get_oggi())
        progetti = get_progetti_dipendente(d["id"])
        result.append({
            "id": d["id"],
            "nome": d["nome"],
            "profilo": d["profilo"],
            "ore_sett": int(d["ore_sett"]),
            "competenze": d["competenze"],
            "carico_corrente": float(carico),
            "saturazione_pct": round(carico / d["ore_sett"] * 100),
            "progetti_attivi": progetti,
            "n_task_attivi": len(_TASKS()[
                (_TASKS()["dipendente_id"] == d["id"]) &
                (_TASKS()["stato"].isin(["In corso", "Da iniziare"]))
            ]),
        })
    return result


@app.get("/api/dipendenti/{dip_id}")
def dettaglio_dipendente(
    dip_id: str, 
    current_user: Utente = Depends(get_current_user),
):
    if current_user.ruolo_app != "manager" and dip_id != current_user.dipendente_id:
        raise HTTPException(403, "Puoi vedere solo il tuo profilo")
    try:
        d = get_dipendente(dip_id)
    except (IndexError, KeyError):
        raise HTTPException(404, "Dipendente non trovato")

    carico = carico_settimanale_dipendente(dip_id, get_oggi())
    progetti = get_progetti_dipendente(dip_id)
    tasks = _TASKS()[
        (_TASKS()["dipendente_id"] == dip_id) &
        (_TASKS()["stato"].isin(["In corso", "Da iniziare"]))
    ]

    return {
        "id": d["id"],
        "nome": d["nome"],
        "profilo": d["profilo"],
        "ore_sett": int(d["ore_sett"]),
        "competenze": d["competenze"],
        "carico_corrente": float(carico),
        "saturazione_pct": round(carico / d["ore_sett"] * 100),
        "progetti_attivi": progetti,
        "tasks": [
            {
                "id": t["id"],
                "nome": t["nome"],
                "progetto": _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]].iloc[0]["nome"],
                "fase": t["fase"],
                "stato": t["stato"],
                "ore_stimate": int(t["ore_stimate"]),
                "data_inizio": t["data_inizio"].isoformat(),
                "data_fine": t["data_fine"].isoformat(),
            }
            for _, t in tasks.iterrows()
        ],
    }


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: _PROGETTI()
# ══════════════════════════════════════════════════════════════════════

@app.get("/api/progetti")
def lista_progetti(_: Utente = Depends(require_manager)):
    result = []
    for _, p in _PROGETTI().iterrows():
        ore_cons = ore_consuntivate_progetto(p["id"])
        tasso = tasso_compilazione_progetto(p["id"])
        tasks_proj = _TASKS()[_TASKS()["progetto_id"] == p["id"]]
        completati = len(tasks_proj[tasks_proj["stato"] == "Completato"])

        result.append({
            "id": p["id"],
            "nome": p["nome"],
            "cliente": p["cliente"],
            "stato": p["stato"],
            "data_inizio": p["data_inizio"].isoformat(),
            "data_fine": p["data_fine"].isoformat(),
            "budget_ore": int(p["budget_ore"]),
            "valore_contratto": float(p["valore_contratto"]),
            "fase_corrente": p["fase_corrente"],
            "ore_consuntivate": float(ore_cons),
            "tasso_compilazione": round(tasso, 1),
            "task_completati": completati,
            "task_totali": len(tasks_proj),
        })
    return result


@app.get("/api/economia/margini")
def economia_margini(_: Utente = Depends(require_manager)):
    """Calcola costi e margini per ogni progetto."""
    result = []
    for _, p in _PROGETTI().iterrows():
        if p["id"] == "P010":
            continue

        tasks_proj = _TASKS()[_TASKS()["progetto_id"] == p["id"]]
        task_ids = tasks_proj["id"].tolist()
        cons_proj = _CONSUNTIVI()[_CONSUNTIVI()["task_id"].isin(task_ids)]

        costo_totale = 0.0
        ore_totali = 0.0
        costi_per_persona = {}

        for _, c in cons_proj.iterrows():
            if c["ore_dichiarate"] <= 0:
                continue
            did = c["dipendente_id"]
            dip = get_dipendente(did)
            costo_ora = float(dip.get("costo_ora", 0)) if dip is not None else 0
            costo_riga = c["ore_dichiarate"] * costo_ora
            costo_totale += costo_riga
            ore_totali += c["ore_dichiarate"]

            if did not in costi_per_persona:
                costi_per_persona[did] = {
                    "nome": dip["nome"], "profilo": dip["profilo"],
                    "costo_ora": costo_ora, "ore": 0, "costo": 0,
                }
            costi_per_persona[did]["ore"] += c["ore_dichiarate"]
            costi_per_persona[did]["costo"] += costo_riga

        valore = float(p["valore_contratto"])
        margine = valore - costo_totale
        margine_pct = (margine / valore * 100) if valore > 0 else 0

        budget_ore = int(p["budget_ore"]) if p["budget_ore"] else 0
        if ore_totali > 0 and budget_ore > 0:
            costo_medio_ora = costo_totale / ore_totali
            costo_stimato_totale = costo_medio_ora * budget_ore
            margine_stimato = valore - costo_stimato_totale
            margine_stimato_pct = (margine_stimato / valore * 100) if valore > 0 else 0
        else:
            costo_stimato_totale = 0
            margine_stimato = valore
            margine_stimato_pct = 100

        result.append({
            "progetto_id": p["id"], "nome": p["nome"],
            "cliente": p["cliente"], "stato": p["stato"],
            "valore_contratto": valore, "budget_ore": budget_ore,
            "ore_consuntivate": round(ore_totali, 1),
            "costo_effettivo": round(costo_totale, 2),
            "margine_attuale": round(margine, 2),
            "margine_pct": round(margine_pct, 1),
            "costo_medio_ora": round(costo_totale / ore_totali, 2) if ore_totali > 0 else 0,
            "costo_stimato_completamento": round(costo_stimato_totale, 2),
            "margine_stimato": round(margine_stimato, 2),
            "margine_stimato_pct": round(margine_stimato_pct, 1),
            "dettaglio_persone": sorted(
                costi_per_persona.values(), key=lambda x: x["costo"], reverse=True
            ),
        })

    return sorted(result, key=lambda x: x["margine_pct"])


@app.get("/api/risorse/suggerisci-bilanciamento")
def suggerisci_bilanciamento(_: Utente = Depends(require_manager)):
    """Analizza le saturazioni e propone redistribuzioni per bilanciare il carico."""
    from datetime import datetime as dt

    oggi = dt.now()
    
    # Calcola saturazione per tutti
    persone = []
    for _, d in _DIPENDENTI().iterrows():
        carico = carico_settimanale_dipendente(d["id"], oggi)
        sat = round(carico / d["ore_sett"] * 100)
        
        # Trova i task attivi di questa persona
        tasks_attivi = _TASKS()[
            (_TASKS()["dipendente_id"] == d["id"]) &
            (_TASKS()["stato"].isin(["In corso", "Da iniziare"])) &
            (_TASKS()["progetto_id"] != "P010")  # escludi attività interne
        ]
        
        task_list = []
        for _, t in tasks_attivi.iterrows():
            # Calcola le ore settimanali di questo task
            durata_giorni = max(1, (t["data_fine"] - t["data_inizio"]).days)
            durata_sett = max(1, durata_giorni / 7)
            ore_sett_task = round(t["ore_stimate"] / durata_sett, 1)
            
            proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]]
            proj_nome = proj.iloc[0]["nome"] if len(proj) > 0 else "?"
            
            task_list.append({
                "task_id": t["id"],
                "task_nome": t["nome"],
                "progetto_id": t["progetto_id"],
                "progetto_nome": proj_nome,
                "profilo_richiesto": t["profilo_richiesto"],
                "ore_stimate": int(t["ore_stimate"]),
                "ore_sett": ore_sett_task,
                "stato": t["stato"],
            })
        
        persone.append({
            "id": d["id"],
            "nome": d["nome"],
            "profilo": d["profilo"],
            "competenze": d["competenze"] if isinstance(d["competenze"], list) else [],
            "ore_sett": int(d["ore_sett"]),
            "carico": float(carico),
            "saturazione": sat,
            "task_attivi": task_list,
        })
    
    # Trova sovraccarichi e sottoutilizzati
    sovraccarichi = [p for p in persone if p["saturazione"] > 100]
    sottoutilizzati = [p for p in persone if p["saturazione"] < 90]
    
    # Genera proposte
    proposte = []
    
    for sov in sorted(sovraccarichi, key=lambda x: -x["saturazione"]):
        eccesso = sov["saturazione"] - 100
        
        for task in sorted(sov["task_attivi"], key=lambda t: t["ore_sett"]):
            # Per ogni task del sovraccarico, cerca candidati
            profilo = task["profilo_richiesto"]
            
            candidati = []
            for sotto in sottoutilizzati:
                if sotto["id"] == sov["id"]:
                    continue
                # Match profilo o competenza
                if sotto["profilo"] == profilo or profilo in sotto["competenze"]:
                    spazio_disponibile = sotto["ore_sett"] - sotto["carico"]
                    if spazio_disponibile >= task["ore_sett"] * 0.5:  # almeno metà delle ore
                        nuova_sat_sov = round((sov["carico"] - task["ore_sett"]) / sov["ore_sett"] * 100)
                        nuova_sat_sotto = round((sotto["carico"] + task["ore_sett"]) / sotto["ore_sett"] * 100)
                        
                        candidati.append({
                            "candidato_id": sotto["id"],
                            "candidato_nome": sotto["nome"],
                            "candidato_profilo": sotto["profilo"],
                            "candidato_saturazione_attuale": sotto["saturazione"],
                            "candidato_saturazione_dopo": nuova_sat_sotto,
                            "spazio_disponibile_h": round(spazio_disponibile, 1),
                        })
            
            if candidati:
                # Ordina candidati: preferisci chi ha più spazio e chi arriva più vicino al 100%
                candidati.sort(key=lambda c: abs(c["candidato_saturazione_dopo"] - 100))
                migliore = candidati[0]
                
                proposte.append({
                    "tipo": "riassegnazione",
                    "priorita": "alta" if sov["saturazione"] > 125 else "media",
                    "da_persona": sov["nome"],
                    "da_persona_id": sov["id"],
                    "da_saturazione": sov["saturazione"],
                    "da_saturazione_dopo": round((sov["carico"] - task["ore_sett"]) / sov["ore_sett"] * 100),
                    "task_id": task["task_id"],
                    "task_nome": task["task_nome"],
                    "progetto": task["progetto_nome"],
                    "ore_sett_task": task["ore_sett"],
                    "profilo_richiesto": profilo,
                    "candidato_migliore": migliore,
                    "altri_candidati": candidati[1:3],  # max 2 alternative
                })
    
    # Ordina: priorità alta prima, poi per eccesso saturazione
    proposte.sort(key=lambda p: (0 if p["priorita"] == "alta" else 1, -p["da_saturazione"]))
    
    return {
        "n_sovraccarichi": len(sovraccarichi),
        "n_sottoutilizzati": len(sottoutilizzati),
        "proposte": proposte,
        "riepilogo": [
            {"nome": p["nome"], "profilo": p["profilo"], "saturazione": p["saturazione"]}
            for p in sorted(persone, key=lambda x: -x["saturazione"])
            if p["saturazione"] > 0
        ],
    }


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT CONFIGURAZIONE — CRUD per pagina admin
# ══════════════════════════════════════════════════════════════════════


# ── RUOLI ──────────────────────────────────────────────────────────────

@app.get("/api/config/ruoli")
def lista_ruoli(_: Utente = Depends(require_manager)):
    session = get_session()
    ruoli = session.query(Ruolo).filter(Ruolo.attivo == True).order_by(Ruolo.nome).all()
    result = [{"id": r.id, "nome": r.nome, "descrizione": r.descrizione or ""} for r in ruoli]
    session.close()
    return result


class RuoloRequest(BaseModel):
    nome: str
    descrizione: str = ""


@app.post("/api/config/ruoli")
def crea_ruolo(req: RuoloRequest, _: Utente = Depends(require_manager)):
    session = get_session()
    existing = session.query(Ruolo).filter(Ruolo.nome == req.nome).first()
    if existing:
        session.close()
        raise HTTPException(400, f"Ruolo '{req.nome}' esiste già")
    ruolo = Ruolo(nome=req.nome, descrizione=req.descrizione)
    session.add(ruolo)
    session.commit()
    result = {"id": ruolo.id, "nome": ruolo.nome}
    session.close()
    return result


@app.patch("/api/config/ruoli/{ruolo_id}")
def modifica_ruolo(ruolo_id: int, req: RuoloRequest, _: Utente = Depends(require_manager)):
    session = get_session()
    ruolo = session.query(Ruolo).filter(Ruolo.id == ruolo_id).first()
    if not ruolo:
        session.close()
        raise HTTPException(404, "Ruolo non trovato")
    ruolo.nome = req.nome
    if req.descrizione:
        ruolo.descrizione = req.descrizione
    session.commit()
    session.close()
    return {"ok": True}


@app.delete("/api/config/ruoli/{ruolo_id}")
def elimina_ruolo(ruolo_id: int, _: Utente = Depends(require_manager)):
    session = get_session()
    ruolo = session.query(Ruolo).filter(Ruolo.id == ruolo_id).first()
    if not ruolo:
        session.close()
        raise HTTPException(404, "Ruolo non trovato")
    ruolo.attivo = False
    session.commit()
    session.close()
    return {"ok": True}


# ── COMPETENZE ─────────────────────────────────────────────────────────

@app.get("/api/config/competenze")
def lista_competenze(_: Utente = Depends(require_manager)):
    session = get_session()
    comps = session.query(Competenza).filter(Competenza.attivo == True).order_by(Competenza.nome).all()
    result = [{"id": c.id, "nome": c.nome} for c in comps]
    session.close()
    return result


class CompetenzaRequest(BaseModel):
    nome: str


@app.post("/api/config/competenze")
def crea_competenza(req: CompetenzaRequest, _: Utente = Depends(require_manager)):
    session = get_session()
    existing = session.query(Competenza).filter(Competenza.nome == req.nome).first()
    if existing:
        session.close()
        raise HTTPException(400, f"Competenza '{req.nome}' esiste già")
    comp = Competenza(nome=req.nome)
    session.add(comp)
    session.commit()
    result = {"id": comp.id, "nome": comp.nome}
    session.close()
    return result


@app.delete("/api/config/competenze/{comp_id}")
def elimina_competenza(comp_id: int, _: Utente = Depends(require_manager)):
    session = get_session()
    comp = session.query(Competenza).filter(Competenza.id == comp_id).first()
    if not comp:
        session.close()
        raise HTTPException(404, "Competenza non trovata")
    comp.attivo = False
    session.commit()
    session.close()
    return {"ok": True}


# ── DIPENDENTI (CRUD arricchito) ───────────────────────────────────────

@app.get("/api/config/dipendenti")
def lista_dipendenti_config(_: Utente = Depends(require_manager)):
    """Dipendenti con ruolo e competenze per la pagina Configurazione."""
    session = get_session()
    dips = session.query(Dipendente).filter(Dipendente.attivo == True).order_by(Dipendente.nome).all()
    result = []
    for d in dips:
        comps = session.query(Competenza.nome).join(DipendentiCompetenze).filter(
            DipendentiCompetenze.dipendente_id == d.id
        ).all()
        result.append({
            "id": d.id,
            "nome": d.nome,
            "profilo": d.profilo,
            "ruolo_id": d.ruolo_id,
            "ore_sett": d.ore_sett,
            "costo_ora": d.costo_ora,
            "email": d.email or "",
            "sede": d.sede or "",
            "competenze": [c[0] for c in comps],
        })
    session.close()
    return result


class DipendenteCfgRequest(BaseModel):
    nome: str
    profilo: str
    ruolo_id: int | None = None
    ore_sett: int = 40
    costo_ora: float | None = None
    email: str = ""
    sede: str = ""
    competenze: list[str] = []


@app.post("/api/config/dipendenti")
def crea_dipendente(req: DipendenteCfgRequest, _: Utente = Depends(require_manager)):
    session = get_session()
    # Genera prossimo ID
    from sqlalchemy import func
    max_id = session.query(func.max(Dipendente.id)).scalar()
    if max_id and max_id.startswith("D") and max_id[1:].isdigit():
        next_num = int(max_id[1:]) + 1
    else:
        next_num = 1
    new_id = f"D{next_num:03d}"

    dip = Dipendente(
        id=new_id, nome=req.nome, profilo=req.profilo,
        ruolo_id=req.ruolo_id, ore_sett=req.ore_sett,
        costo_ora=req.costo_ora, email=req.email, sede=req.sede,
        competenze=req.competenze,
    )
    session.add(dip)
    session.flush()

    # Associa competenze M2M
    for comp_nome in req.competenze:
        comp = session.query(Competenza).filter(Competenza.nome == comp_nome).first()
        if comp:
            session.add(DipendentiCompetenze(dipendente_id=new_id, competenza_id=comp.id))

    session.commit()
    session.close()
    return {"id": new_id, "nome": req.nome}


@app.patch("/api/config/dipendenti/{dip_id}")
def modifica_dipendente(dip_id: str, req: DipendenteCfgRequest, _: Utente = Depends(require_manager)):
    session = get_session()
    dip = session.query(Dipendente).filter(Dipendente.id == dip_id).first()
    if not dip:
        session.close()
        raise HTTPException(404, "Dipendente non trovato")

    dip.nome = req.nome
    dip.profilo = req.profilo
    dip.ruolo_id = req.ruolo_id
    dip.ore_sett = req.ore_sett
    dip.costo_ora = req.costo_ora
    dip.email = req.email
    dip.sede = req.sede
    dip.competenze = req.competenze

    # Aggiorna competenze M2M
    session.query(DipendentiCompetenze).filter(
        DipendentiCompetenze.dipendente_id == dip_id
    ).delete()
    for comp_nome in req.competenze:
        comp = session.query(Competenza).filter(Competenza.nome == comp_nome).first()
        if comp:
            session.add(DipendentiCompetenze(dipendente_id=dip_id, competenza_id=comp.id))

    session.commit()
    session.close()
    return {"ok": True}


@app.delete("/api/config/dipendenti/{dip_id}")
def elimina_dipendente(dip_id: str, _: Utente = Depends(require_manager)):
    session = get_session()
    dip = session.query(Dipendente).filter(Dipendente.id == dip_id).first()
    if not dip:
        session.close()
        raise HTTPException(404, "Dipendente non trovato")
    dip.attivo = False
    session.commit()
    session.close()
    return {"ok": True}


# ── FASI STANDARD (template) ──────────────────────────────────────────

@app.get("/api/config/fasi-standard")
def lista_fasi_standard(_: Utente = Depends(require_manager)):
    """Restituisce i template raggruppati per nome template."""
    session = get_session()
    fasi = session.query(FaseStandard).order_by(FaseStandard.template_nome, FaseStandard.ordine).all()
    # Raggruppa per template
    templates = {}
    for f in fasi:
        if f.template_nome not in templates:
            templates[f.template_nome] = []
        templates[f.template_nome].append({
            "id": f.id,
            "fase_nome": f.fase_nome,
            "ordine": f.ordine,
            "percentuale_ore": f.percentuale_ore,
        })
    session.close()
    return templates


class FaseStandardRequest(BaseModel):
    template_nome: str
    fase_nome: str
    ordine: int = 1
    percentuale_ore: float | None = None


@app.post("/api/config/fasi-standard")
def crea_fase_standard(req: FaseStandardRequest, _: Utente = Depends(require_manager)):
    session = get_session()
    fs = FaseStandard(
        template_nome=req.template_nome,
        fase_nome=req.fase_nome,
        ordine=req.ordine,
        percentuale_ore=req.percentuale_ore,
    )
    session.add(fs)
    session.commit()
    result = {"id": fs.id, "template_nome": req.template_nome, "fase_nome": req.fase_nome}
    session.close()
    return result


@app.delete("/api/config/fasi-standard/{fs_id}")
def elimina_fase_standard(fs_id: int, _: Utente = Depends(require_manager)):
    session = get_session()
    fs = session.query(FaseStandard).filter(FaseStandard.id == fs_id).first()
    if not fs:
        session.close()
        raise HTTPException(404, "Fase standard non trovata")
    session.delete(fs)
    session.commit()
    session.close()
    return {"ok": True}


@app.get("/api/config/fasi-catalogo")
def lista_fasi_catalogo(_: Utente = Depends(require_manager)):
    """Lista piatta di fasi disponibili per la pianificazione."""
    session = get_session()
    fasi = session.query(FaseStandard).filter(
        FaseStandard.template_nome == "_catalogo"
    ).order_by(FaseStandard.ordine).all()
    
    # Se non ci sono fasi nel catalogo, carica quelle uniche dai template esistenti
    if not fasi:
        fasi_template = session.query(FaseStandard).order_by(FaseStandard.ordine).all()
        nomi_visti = set()
        result = []
        for f in fasi_template:
            if f.fase_nome not in nomi_visti:
                nomi_visti.add(f.fase_nome)
                result.append({"id": f.id, "nome": f.fase_nome, "ordine": f.ordine})
        session.close()
        return result
    
    result = [{"id": f.id, "nome": f.fase_nome, "ordine": f.ordine} for f in fasi]
    session.close()
    return result
 
 
class FaseCatalogoRequest(BaseModel):
    nome: str
 
 
@app.post("/api/config/fasi-catalogo")
def crea_fase_catalogo(req: FaseCatalogoRequest, _: Utente = Depends(require_manager)):
    """Aggiunge una fase al catalogo."""
    session = get_session()
    # Controlla duplicati
    existing = session.query(FaseStandard).filter(
        FaseStandard.template_nome == "_catalogo",
        FaseStandard.fase_nome == req.nome
    ).first()
    if existing:
        session.close()
        raise HTTPException(400, f"Fase '{req.nome}' esiste già")
    
    # Trova prossimo ordine
    from sqlalchemy import func
    max_ordine = session.query(func.max(FaseStandard.ordine)).filter(
        FaseStandard.template_nome == "_catalogo"
    ).scalar() or 0
    
    fs = FaseStandard(
        template_nome="_catalogo",
        fase_nome=req.nome,
        ordine=max_ordine + 1,
    )
    session.add(fs)
    session.commit()
    result = {"id": fs.id, "nome": fs.fase_nome, "ordine": fs.ordine}
    session.close()
    return result
 
 
@app.delete("/api/config/fasi-catalogo/{fase_id}")
def elimina_fase_catalogo(fase_id: int, _: Utente = Depends(require_manager)):
    session = get_session()
    fs = session.query(FaseStandard).filter(FaseStandard.id == fase_id).first()
    if not fs:
        session.close()
        raise HTTPException(404, "Fase non trovata")
    session.delete(fs)
    session.commit()
    session.close()
    return {"ok": True}


@app.post("/api/fasi")
def crea_fase(req: dict, _: Utente = Depends(require_manager)):
    session = get_session()
    fase = Fase(
        progetto_id=req["progetto_id"],
        nome=req["nome"],
        ordine=req.get("ordine", 1),
        data_inizio=req.get("data_inizio"),
        data_fine=req.get("data_fine"),
        ore_vendute=req.get("ore_vendute", 0),
        stato="Da iniziare",
    )
    session.add(fase)
    session.commit()
    result = {"id": fase.id, "nome": fase.nome}
    session.close()
    return result


# ── FASI DI PROGETTO ───────────────────────────────────────────────────

@app.get("/api/fasi/{progetto_id}")
def lista_fasi_progetto(progetto_id: str, _: Utente = Depends(require_manager)):
    """Lista fasi di un progetto con totali ore."""
    session = get_session()
    fasi = session.query(Fase).filter(Fase.progetto_id == progetto_id).order_by(Fase.ordine).all()
    result = []
    for f in fasi:
        # Calcola ore consumate dai consuntivi dei task in questa fase
        tasks_fase = session.query(Task).filter(Task.fase_id == f.id).all()
        task_ids = [t.id for t in tasks_fase]
        ore_consumate = 0
        if task_ids:
            from sqlalchemy import func
            ore_consumate = session.query(func.coalesce(func.sum(Consuntivo.ore_dichiarate), 0)).filter(
                Consuntivo.task_id.in_(task_ids)
            ).scalar()

        result.append({
            "id": f.id,
            "nome": f.nome,
            "ordine": f.ordine,
            "data_inizio": f.data_inizio.isoformat() if f.data_inizio else None,
            "data_fine": f.data_fine.isoformat() if f.data_fine else None,
            "ore_vendute": f.ore_vendute,
            "ore_pianificate": f.ore_pianificate,
            "ore_consumate": float(ore_consumate),
            "ore_rimanenti": (f.ore_vendute or 0) - float(ore_consumate),
            "stato": f.stato,
            "n_task": len(tasks_fase),
            "note": f.note or "",
        })
    session.close()
    return result


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: _TASKS() / GANTT
# ══════════════════════════════════════════════════════════════════════

@app.get("/api/tasks")
def lista_tasks(
    progetto_id: Optional[str] = None, 
    profilo: Optional[str] = None,
    current_user: Utente = Depends(get_current_user),
):
    tasks = _TASKS().copy()
    tasks = tasks[tasks["stato"] != "Eliminato"]
    # Scenario B: user vede solo i propri task, manager vede tutto
    if current_user.ruolo_app != "manager":
        tasks = tasks[tasks["dipendente_id"] == current_user.dipendente_id]
    if progetto_id:
        tasks = tasks[tasks["progetto_id"] == progetto_id]
    if profilo:
        tasks = tasks[tasks["profilo_richiesto"] == profilo]

    result = []
    for _, t in tasks.iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]].iloc[0]
        result.append({
            "id": t["id"],
            "nome": t["nome"],
            "progetto_id": t["progetto_id"],
            "progetto_nome": proj["nome"],
            "fase": t["fase"],
            "stato": t["stato"],
            "ore_stimate": int(t["ore_stimate"]),
            "data_inizio": t["data_inizio"].isoformat(),
            "data_fine": t["data_fine"].isoformat(),
            "profilo_richiesto": t["profilo_richiesto"],
            "dipendente_id": t["dipendente_id"],
            "dipendente_nome": dip["nome"],
            "predecessore": t["predecessore"],
        })
    return result


@app.get("/api/gantt")
def dati_gantt(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Restituisce i dati formattati per un componente GANTT."""
    tasks = _TASKS().copy()
    tasks = tasks[tasks["stato"] != "Eliminato"]
    if progetto_id:
        tasks = tasks[tasks["progetto_id"] == progetto_id]

    result = []
    for _, t in tasks.iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]].iloc[0]
        # Calcola ore consuntivate reali per questo task
        cons_task = _CONSUNTIVI()[_CONSUNTIVI()["task_id"] == t["id"]]
        ore_cons = float(cons_task["ore_dichiarate"].sum()) if len(cons_task) > 0 else 0
        ore_stimate = int(t["ore_stimate"]) if t["ore_stimate"] else 0
 
        # Progress reale: basato su ore consuntivate / ore stimate
        if t["stato"] == "Completato":
            progress = 100
        elif ore_stimate > 0 and ore_cons > 0:
            progress = min(99, round(ore_cons / ore_stimate * 100))  # max 99 se non completato
        else:
            progress = 0
 
        result.append({
            "id": t["id"],
            "name": t["nome"],
            "start": t["data_inizio"].strftime("%Y-%m-%d"),
            "end": t["data_fine"].strftime("%Y-%m-%d"),
            "progress": progress,
            "dependencies": t["predecessore"] if t["predecessore"] else "",
            "predecessor_name": "",  # popolato sotto se c'è predecessore
            "project": proj["nome"],
            "project_id": t["progetto_id"],
            "assignee": dip["nome"],
            "profile": dip["profilo"],
            "status": t["stato"],
            "estimated_hours": ore_stimate,
            "hours_done": round(ore_cons, 1),
        })
 
        # Aggiungi nome predecessore (per il pannello dettaglio)
        if t["predecessore"]:
            pred_row = _TASKS()[_TASKS()["id"] == t["predecessore"]]
            if len(pred_row) > 0:
                result[-1]["predecessor_name"] = pred_row.iloc[0]["nome"]

    return result


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: CARICO RISORSE (heatmap)
# ══════════════════════════════════════════════════════════════════════

@app.get("/api/risorse/carico")
def carico_risorse(
    settimane: int = 12, 
    _: Utente = Depends(require_manager),
):
    """Restituisce la heatmap di saturazione per tutte le risorse."""
    from datetime import timedelta
    result = []
    for _, d in _DIPENDENTI().iterrows():
        settimane_data = []
        for w in range(settimane):
            sett = get_oggi() + timedelta(weeks=w)
            lun = sett - timedelta(days=sett.weekday())
            carico = carico_settimanale_dipendente(d["id"], sett)
            settimane_data.append({
                "settimana": lun.strftime("%Y-%m-%d"),
                "settimana_label": lun.strftime("%d/%m"),
                "ore_assegnate": float(carico),
                "saturazione_pct": min(125, round(carico / d["ore_sett"] * 100)),
            })
        result.append({
            "dipendente_id": d["id"],
            "nome": d["nome"],
            "profilo": d["profilo"],
            "ore_sett": int(d["ore_sett"]),
            "settimane": settimane_data,
        })
    return result


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: CHATBOT CONSUNTIVAZIONE
# ══════════════════════════════════════════════════════════════════════

@app.get("/api/agent/status")
def agent_status(_: Utente = Depends(get_current_user)):
    return {"available": is_agent_available()}


@app.post("/api/agent/chat")
def agent_chat(req: ChatRequest):
    import json as json_mod

    try:
        dip = get_dipendente(req.dipendente_id)
    except (IndexError, KeyError):
        raise HTTPException(404, "Dipendente non trovato")

    tasks_attivi = _TASKS()[
        (_TASKS()["dipendente_id"] == req.dipendente_id) &
        (_TASKS()["stato"].isin(["In corso", "Da iniziare"]))
    ]

    contesto = costruisci_contesto(
        dip_data=dip,
        ore_compilate=req.ore_compilate,
        stati_compilati=req.stati_compilati,
        tasks_attivi=tasks_attivi,
        ore_assenza=req.ore_assenza,
        tipo_assenza=req.tipo_assenza,
        nota_assenza=req.nota_assenza,
        spese=req.spese if req.spese else None,
        ore_contrattuali=dip["ore_sett"],
    )

    model = init_gemini()
    risposta = chiedi_agente(model, contesto, req.messaggio, chat_history=req.chat_history)

    # ── Parsa blocchi strutturati dalla risposta ──
    # Sia [MAPPATURA_ORE] che [SEGNALAZIONE] vengono estratti
    # e rimossi dalla risposta visibile al dipendente.

    risposta_pulita = risposta
    segnalazione = None
    mappatura_ore = None

    # 1) Parsa [MAPPATURA_ORE]
    if "[MAPPATURA_ORE]" in risposta_pulita:
        idx = risposta_pulita.index("[MAPPATURA_ORE]")
        mappatura_raw = risposta_pulita[idx + len("[MAPPATURA_ORE]"):]
        risposta_pulita = risposta_pulita[:idx].strip()

        # Se c'è anche una segnalazione dopo la mappatura, separale
        if "[SEGNALAZIONE]" in mappatura_raw:
            segn_idx = mappatura_raw.index("[SEGNALAZIONE]")
            segn_part = mappatura_raw[segn_idx:]
            mappatura_raw = mappatura_raw[:segn_idx]
            # La segnalazione la processiamo dopo
            risposta_pulita_con_segn = risposta_pulita + "\n" + segn_part
        else:
            risposta_pulita_con_segn = None

        # Estrai il JSON dalla mappatura
        mappatura_raw = mappatura_raw.strip()
        # Rimuovi eventuali backtick markdown
        if mappatura_raw.startswith("```json"):
            mappatura_raw = mappatura_raw[7:]
        if mappatura_raw.startswith("```"):
            mappatura_raw = mappatura_raw[3:]
        if mappatura_raw.endswith("```"):
            mappatura_raw = mappatura_raw[:-3]
        mappatura_raw = mappatura_raw.strip()

        try:
            mappatura_ore = json_mod.loads(mappatura_raw)
        except json_mod.JSONDecodeError:
            # Se non riesce a parsare, ignora la mappatura
            mappatura_ore = None

        # Se c'era una segnalazione dopo la mappatura, rimettila per il parsing
        if risposta_pulita_con_segn:
            risposta_pulita = risposta_pulita  # resta pulita
            # Processiamo la segnalazione dalla parte estratta
            risposta_per_segn = segn_part
        else:
            risposta_per_segn = ""
    else:
        risposta_per_segn = risposta_pulita

    # 2) Parsa [SEGNALAZIONE]
    testo_da_parsare = risposta_per_segn if "[SEGNALAZIONE]" in risposta_per_segn else risposta_pulita
    if "[SEGNALAZIONE]" in testo_da_parsare:
        idx = testo_da_parsare.index("[SEGNALAZIONE]")
        segnalazione_raw = testo_da_parsare[idx:]

        # Rimuovi dalla risposta visibile (se non già rimossa dalla mappatura)
        if "[SEGNALAZIONE]" in risposta_pulita:
            risposta_pulita = risposta_pulita[:risposta_pulita.index("[SEGNALAZIONE]")].strip()

        segnalazione = segnalazione_raw

        # Parsa tipo, dettaglio, priorità dalla segnalazione
        segn_tipo = "generico"
        segn_priorita = "media"
        segn_dettaglio = segnalazione_raw
        for line in segnalazione_raw.split("\n"):
            line_lower = line.strip().lower()
            if "tipo:" in line_lower:
                segn_tipo = line.split(":", 1)[1].strip().lower().replace(" ", "_")
            elif "priorit" in line_lower:
                val = line.split(":", 1)[1].strip().lower()
                if "alta" in val: segn_priorita = "alta"
                elif "bassa" in val: segn_priorita = "bassa"
            elif "dettaglio:" in line_lower or "descrizione:" in line_lower:
                segn_dettaglio = line.split(":", 1)[1].strip()

        if PERSISTENT_MODE:
            aggiungi_segnalazione(segn_tipo, segn_priorita, req.dipendente_id, segn_dettaglio)
        else:
            SEGNALAZIONI_STORE.append({
                "id": _next_segn_id(),
                "tipo": segn_tipo,
                "priorita": segn_priorita,
                "dipendente_id": req.dipendente_id,
                "dipendente": dip["nome"],
                "dettaglio": segn_dettaglio,
                "timestamp": get_oggi().strftime("%Y-%m-%d %H:%M"),
            })

    return {
        "risposta": risposta_pulita,
        "segnalazione": segnalazione,
        "mappatura_ore": mappatura_ore,
        "contesto_carico": contesto.get("carico_complessivo"),
    }


@app.get("/api/segnalazioni")
def lista_segnalazioni(_: Utente = Depends(require_manager)):
    """Restituisce tutte le segnalazioni raccolte."""
    if PERSISTENT_MODE:
        return get_segnalazioni()
    return SEGNALAZIONI_STORE


# ══════════════════════════════════════════════════════════════
# AGGIUNGI QUESTO ENDPOINT in main.py (vicino agli altri endpoint consuntivi)
# ══════════════════════════════════════════════════════════════
 
@app.get("/api/consuntivi/settimana")
def consuntivi_settimana_corrente():
    """Restituisce i consuntivi della settimana corrente per tutti i dipendenti."""
    from datetime import timedelta, datetime as dt_now
    lun = dt_now.now() - timedelta(days=dt_now.now().weekday())
    lun_date = lun.date() if hasattr(lun, 'date') else lun
 
    consuntivi = _CONSUNTIVI()
    if consuntivi.empty:
        return []
 
    # Filtra per settimana corrente
    ven_date = lun_date + timedelta(days=6)
    cons_sett = consuntivi[consuntivi["settimana"].apply(
    lambda x: lun_date <= (x.date() if hasattr(x, 'date') else x) <= ven_date
)]
 
    # Raggruppa per dipendente
    risultato = []
    for did in cons_sett["dipendente_id"].unique():
        try:
            dip = get_dipendente(did)
        except:
            continue
        cons_dip = cons_sett[cons_sett["dipendente_id"] == did]
        ore_per_task = []
        totale = 0
        for _, c in cons_dip.iterrows():
            if c["ore_dichiarate"] > 0:
                # Trova nome task e progetto
                task_row = _TASKS()[_TASKS()["id"] == c["task_id"]]
                if not task_row.empty:
                    t = task_row.iloc[0]
                    proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]]
                    proj_nome = proj.iloc[0]["nome"] if not proj.empty else "?"
                    ore_per_task.append({
                        "task_nome": t["nome"],
                        "progetto": proj_nome,
                        "ore": float(c["ore_dichiarate"]),
                    })
                    totale += float(c["ore_dichiarate"])
 
        if ore_per_task:
            risultato.append({
                "dipendente_id": did,
                "nome": dip["nome"],
                "profilo": dip["profilo"],
                "ore_contrattuali": int(dip["ore_sett"]),
                "totale_ore": round(totale, 1),
                "ore_per_task": ore_per_task,
                "compilato": True,
            })
 
    # Aggiungi dipendenti che NON hanno compilato
    for _, d in _DIPENDENTI().iterrows():
        if d["id"] not in [r["dipendente_id"] for r in risultato]:
            n_task = len(_TASKS()[
                (_TASKS()["dipendente_id"] == d["id"]) &
                (_TASKS()["stato"].isin(["In corso", "Da iniziare"]))
            ])
            if n_task > 0:
                risultato.append({
                    "dipendente_id": d["id"],
                    "nome": d["nome"],
                    "profilo": d["profilo"],
                    "ore_contrattuali": int(d["ore_sett"]),
                    "totale_ore": 0,
                    "ore_per_task": [],
                    "compilato": False,
                })
 
    return sorted(risultato, key=lambda x: (-x["compilato"], x["nome"]))


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: SIMULAZIONI
# ══════════════════════════════════════════════════════════════════════

@app.post("/api/simulazione/ritardo")
def simula_ritardo(req: SimulaRitardoRequest):
    from datetime import timedelta

    tasks_sim = _TASKS().copy()
    task_sel = tasks_sim[tasks_sim["id"] == req.task_id]
    if len(task_sel) == 0:
        raise HTTPException(404, "Task non trovato")

    task_sel = task_sel.iloc[0]
    nuova_fine = task_sel["data_fine"] + timedelta(days=req.giorni_ritardo)
    tasks_sim.loc[tasks_sim["id"] == req.task_id, "data_fine"] = nuova_fine

    impatti = []

    def propaga(df, tid, nuova_fine_pred):
        successori = df[df["predecessore"] == tid]
        for idx, succ in successori.iterrows():
            durata = (succ["data_fine"] - succ["data_inizio"]).days
            nuovo_inizio = nuova_fine_pred + timedelta(days=1)
            nuova_fine_s = nuovo_inizio + timedelta(days=durata)
            df.loc[idx, "data_inizio"] = nuovo_inizio
            df.loc[idx, "data_fine"] = nuova_fine_s

            dip = get_dipendente(succ["dipendente_id"])
            carico = carico_settimanale_dipendente(succ["dipendente_id"], nuovo_inizio)

            impatti.append({
                "task_id": succ["id"],
                "task_nome": succ["nome"],
                "dipendente": dip["nome"],
                "nuovo_inizio": nuovo_inizio.isoformat(),
                "nuova_fine": nuova_fine_s.isoformat(),
                "sovraccarico": bool(carico > dip["ore_sett"]),
                "carico": float(carico),
                "capacita": int(dip["ore_sett"]),
            })
            propaga(df, succ["id"], nuova_fine_s)
        return df

    tasks_sim = propaga(tasks_sim, req.task_id, nuova_fine)

    # Restituisci GANTT aggiornato
    gantt_sim = []
    for _, t in tasks_sim.iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]].iloc[0]
        gantt_sim.append({
            "id": t["id"],
            "name": t["nome"],
            "start": t["data_inizio"].strftime("%Y-%m-%d"),
            "end": t["data_fine"].strftime("%Y-%m-%d"),
            "progress": 100 if t["stato"] == "Completato" else 50 if t["stato"] == "In corso" else 0,
            "dependencies": t["predecessore"] if t["predecessore"] else "",
            "project": proj["nome"],
            "assignee": dip["nome"],
            "status": t["stato"],
        })

    return {
        "task_ritardato": task_sel["nome"],
        "nuova_fine": nuova_fine.isoformat(),
        "impatti": impatti,
        "gantt": gantt_sim,
    }


@app.post("/api/simulazione/ritardo-multiplo")
def simula_ritardo_multiplo(req: SimulaRitardoMultiploRequest):
    """Simula il ritardo di più task contemporaneamente con propagazione a cascata."""
    from datetime import timedelta

    if not req.ritardi:
        raise HTTPException(400, "Nessun ritardo specificato")

    tasks_sim = _TASKS().copy()
    impatti = []
    task_ritardati = []

    def propaga(df, tid, nuova_fine_pred, already_shifted):
        successori = df[df["predecessore"] == tid]
        for idx, succ in successori.iterrows():
            if succ["id"] in already_shifted:
                continue
            durata = (succ["data_fine"] - succ["data_inizio"]).days
            nuovo_inizio = nuova_fine_pred + timedelta(days=1)
            nuova_fine_s = nuovo_inizio + timedelta(days=durata)
            df.loc[idx, "data_inizio"] = nuovo_inizio
            df.loc[idx, "data_fine"] = nuova_fine_s
            already_shifted.add(succ["id"])

            dip = get_dipendente(succ["dipendente_id"])
            carico = carico_settimanale_dipendente(succ["dipendente_id"], nuovo_inizio)
            proj = _PROGETTI()[_PROGETTI()["id"] == succ["progetto_id"]].iloc[0]

            impatti.append({
                "task_id": succ["id"],
                "task_nome": succ["nome"],
                "progetto": proj["nome"],
                "dipendente": dip["nome"],
                "nuovo_inizio": nuovo_inizio.isoformat(),
                "nuova_fine": nuova_fine_s.isoformat(),
                "sovraccarico": bool(carico > dip["ore_sett"]),
                "carico": float(carico),
                "capacita": int(dip["ore_sett"]),
            })
            propaga(df, succ["id"], nuova_fine_s, already_shifted)
        return df

    already_shifted = set()

    # Applica tutti i ritardi diretti prima
    for r in req.ritardi:
        task_sel = tasks_sim[tasks_sim["id"] == r.task_id]
        if len(task_sel) == 0:
            continue
        task_row = task_sel.iloc[0]
        proj = _PROGETTI()[_PROGETTI()["id"] == task_row["progetto_id"]].iloc[0]
        nuova_fine = task_row["data_fine"] + timedelta(days=r.giorni_ritardo)
        tasks_sim.loc[tasks_sim["id"] == r.task_id, "data_fine"] = nuova_fine
        already_shifted.add(r.task_id)
        task_ritardati.append({
            "task_id": r.task_id,
            "task_nome": task_row["nome"],
            "progetto": proj["nome"],
            "giorni": r.giorni_ritardo,
            "nuova_fine": nuova_fine.isoformat(),
        })

    # Poi propaga a cascata per ognuno
    for r in req.ritardi:
        nuova_fine = tasks_sim[tasks_sim["id"] == r.task_id].iloc[0]["data_fine"]
        tasks_sim = propaga(tasks_sim, r.task_id, nuova_fine, already_shifted)

    # GANTT originale (per confronto prima/dopo)
    gantt_prima = []
    for _, t in _TASKS().iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]].iloc[0]
        gantt_prima.append({
            "id": t["id"],
            "name": t["nome"],
            "start": t["data_inizio"].strftime("%Y-%m-%d"),
            "end": t["data_fine"].strftime("%Y-%m-%d"),
            "progress": 100 if t["stato"] == "Completato" else 50 if t["stato"] == "In corso" else 0,
            "dependencies": t["predecessore"] if t["predecessore"] else "",
            "project": proj["nome"],
            "assignee": dip["nome"],
            "profile": dip["profilo"],
            "status": t["stato"],
            "estimated_hours": int(t["ore_stimate"]),
        })

    # GANTT simulato
    gantt_dopo = []
    changed_ids = already_shifted
    for _, t in tasks_sim.iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]].iloc[0]
        gantt_dopo.append({
            "id": t["id"],
            "name": t["nome"],
            "start": t["data_inizio"].strftime("%Y-%m-%d"),
            "end": t["data_fine"].strftime("%Y-%m-%d"),
            "progress": 100 if t["stato"] == "Completato" else 50 if t["stato"] == "In corso" else 0,
            "dependencies": t["predecessore"] if t["predecessore"] else "",
            "project": proj["nome"],
            "assignee": dip["nome"],
            "profile": dip["profilo"],
            "status": t["stato"],
            "estimated_hours": int(t["ore_stimate"]),
            "changed": bool(t["id"] in changed_ids),  # flag per evidenziare
        })

    return {
        "task_ritardati": task_ritardati,
        "impatti": impatti,
        "gantt_prima": gantt_prima,
        "gantt_dopo": gantt_dopo,
        "changed_ids": list(changed_ids),
    }


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: AGENTE ANALISI GANTT
# ══════════════════════════════════════════════════════════════════════

class AnalisiRequest(BaseModel):
    segnalazione_tipo: str
    segnalazione_dettaglio: str
    dipendente_id: str
    priorita: str = "media"


@app.post("/api/agent/analisi-gantt")
def analisi_gantt(req: AnalisiRequest):
    """Riceve una segnalazione e restituisce proposte di redistribuzione."""
    import json as json_mod

    # Costruisci contesto completo per l'agente
    try:
        dip_segnalante = get_dipendente(req.dipendente_id)
    except (IndexError, KeyError):
        raise HTTPException(400, f"Dipendente '{req.dipendente_id}' non trovato")

    # Tutti i dipendenti con carico
    dip_contesto = []
    for _, d in _DIPENDENTI().iterrows():
        carico = carico_settimanale_dipendente(d["id"], get_oggi())
        progetti = get_progetti_dipendente(d["id"])
        tasks_attivi = _TASKS()[
            (_TASKS()["dipendente_id"] == d["id"]) &
            (_TASKS()["stato"].isin(["In corso", "Da iniziare"]))
        ]
        dip_contesto.append({
            "id": d["id"],
            "nome": d["nome"],
            "profilo": d["profilo"],
            "ore_sett": int(d["ore_sett"]),
            "carico_corrente": float(carico),
            "saturazione_pct": round(carico / d["ore_sett"] * 100),
            "progetti": progetti,
            "task_attivi": [
                {
                    "id": t["id"],
                    "nome": t["nome"],
                    "progetto": _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]].iloc[0]["nome"],
                    "ore_stimate": int(t["ore_stimate"]),
                    "data_inizio": t["data_inizio"].strftime("%Y-%m-%d"),
                    "data_fine": t["data_fine"].strftime("%Y-%m-%d"),
                    "stato": t["stato"],
                    "profilo_richiesto": t["profilo_richiesto"],
                    "predecessore": t["predecessore"] if t["predecessore"] else None,
                }
                for _, t in tasks_attivi.iterrows()
            ],
        })

    # Tutti i progetti con scadenze
    proj_contesto = []
    for _, p in _PROGETTI().iterrows():
        if p["stato"] in ["In esecuzione", "In bando"]:
            proj_contesto.append({
                "id": p["id"],
                "nome": p["nome"],
                "cliente": p["cliente"],
                "stato": p["stato"],
                "data_fine": p["data_fine"].strftime("%Y-%m-%d"),
                "budget_ore": int(p["budget_ore"]),
            })

    contesto_completo = {
        "segnalazione": {
            "tipo": req.segnalazione_tipo,
            "dettaglio": req.segnalazione_dettaglio,
            "dipendente": dip_segnalante["nome"],
            "dipendente_id": req.dipendente_id,
            "priorita": req.priorita,
        },
        "dipendenti": dip_contesto,
        "progetti": proj_contesto,
        "data_corrente": get_oggi().strftime("%Y-%m-%d"),
    }

    # Chiama Gemini con il prompt specializzato
    model = init_gemini(prompt_file="analisi_segnalazioni.md")
    if model is None:
        raise HTTPException(503, "Agente AI non disponibile")

    contesto_json = json_mod.dumps(contesto_completo, ensure_ascii=False, indent=2)
    prompt = f"Analizza questa segnalazione e proponi redistribuzioni:\n\n```json\n{contesto_json}\n```"

    try:
        chat = model.start_chat(history=[])
        response = chat.send_message(prompt)
        risposta_raw = response.text

        # Tenta di parsare il JSON dalla risposta
        # Rimuovi eventiali backtick markdown
        cleaned = risposta_raw.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()

        try:
            proposte = json_mod.loads(cleaned)
        except json_mod.JSONDecodeError:
            # Se non riesce a parsare, restituisci il testo grezzo
            proposte = {"raw_response": risposta_raw, "parse_error": True}

        return {
            "proposte": proposte,
            "contesto": {
                "segnalazione": contesto_completo["segnalazione"],
                "dipendente_saturazione": next(
                    (d["saturazione_pct"] for d in dip_contesto if d["id"] == req.dipendente_id), 0
                ),
            },
        }

    except Exception as e:
        raise HTTPException(500, f"Errore agente: {str(e)}")


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: ANTEPRIMA IMPATTO + APPLICA MODIFICHE
# ══════════════════════════════════════════════════════════════════════

class AzioneModifica(BaseModel):
    task_id: str
    campo: str           # "data_fine", "data_inizio", "dipendente_id", "ore_stimate", "stato"
    nuovo_valore: str    # tutto come stringa, il backend converte


class NuovoTask(BaseModel):
    nome: str
    fase: str = ""
    ore_stimate: int = 0
    data_inizio: str = ""        # ISO format
    data_fine: str = ""          # ISO format
    profilo_richiesto: str = ""
    dipendente_id: str = ""
    predecessore: str = ""
    stato: str = "Da iniziare"


class AnteprimaRequest(BaseModel):
    """Richiesta di anteprima impatto — non modifica nulla."""
    modifiche: list[AzioneModifica] = []
    nuovi_task: list[NuovoTask] = []
    progetto_id: str = ""        # per nuovi task senza progetto esplicito


class ApplicaRequest(BaseModel):
    """Richiesta di applicazione reale — modifica i dati."""
    modifiche: list[AzioneModifica] = []
    nuovi_task: list[NuovoTask] = []
    progetto_id: str = ""
    cambia_stato_progetto: str = ""  # es. "In esecuzione"


@app.post("/api/task/anteprima-impatto")
def anteprima_impatto(req: AnteprimaRequest, _: Utente = Depends(require_manager)):
    """
    Calcola l'impatto delle modifiche proposte SENZA applicarle.
    Restituisce saturazioni prima/dopo e alert per il management.
    """
    from datetime import datetime as dt

    # Converti modifiche nel formato atteso da calcola_impatto_saturazione
    task_modifiche = []
    for mod in req.modifiche:
        valore = mod.nuovo_valore
        # Converti date da stringa ISO a datetime
        if mod.campo in ("data_inizio", "data_fine"):
            valore = dt.fromisoformat(valore)
        elif mod.campo == "ore_stimate":
            valore = int(valore)
        task_modifiche.append({
            "task_id": mod.task_id,
            "campo": mod.campo,
            "nuovo_valore": valore,
        })

    # Converti nuovi task
    task_nuovi = []
    for nt in req.nuovi_task:
        task_nuovi.append({
            "id": f"PREVIEW_{len(task_nuovi)}",
            "progetto_id": req.progetto_id,
            "nome": nt.nome,
            "fase": nt.fase,
            "ore_stimate": nt.ore_stimate,
            "data_inizio": dt.fromisoformat(nt.data_inizio) if nt.data_inizio else get_oggi(),
            "data_fine": dt.fromisoformat(nt.data_fine) if nt.data_fine else get_oggi(),
            "stato": nt.stato,
            "profilo_richiesto": nt.profilo_richiesto,
            "dipendente_id": nt.dipendente_id,
            "predecessore": nt.predecessore,
        })

    impatto = calcola_impatto_saturazione(task_modifiche, task_nuovi if task_nuovi else None)

    # Aggiungi GANTT simulato dei progetti impattati
    gantt_impattati = {}
    for proj in impatto["progetti_impattati"]:
        tasks_proj = get_tasks_progetto(proj["id"])
        gantt_tasks = []
        for _, t in tasks_proj.iterrows():
            try:
                dip = get_dipendente(t["dipendente_id"])
                gantt_tasks.append({
                    "id": t["id"],
                    "name": t["nome"],
                    "start": t["data_inizio"].strftime("%Y-%m-%d"),
                    "end": t["data_fine"].strftime("%Y-%m-%d"),
                    "progress": 100 if t["stato"] == "Completato" else 50 if t["stato"] == "In corso" else 0,
                    "assignee": dip["nome"],
                    "status": t["stato"],
                })
            except (IndexError, KeyError):
                pass
        gantt_impattati[proj["id"]] = gantt_tasks

    return {
        "impatto": impatto,
        "gantt_progetti_impattati": gantt_impattati,
    }


@app.post("/api/task/applica")
def applica_modifiche(req: ApplicaRequest, _: Utente = Depends(require_manager)):
    """
    Applica le modifiche ai dati reali.
    Usato sia da Analisi e Interventi (bottone Applica)
    sia da Pipeline (Conferma e avvia progetto).
    """
    from datetime import datetime as dt

    risultati = []

    # 1) Applica modifiche a task esistenti
    for mod in req.modifiche:
        valore = mod.nuovo_valore
        if mod.campo in ("data_inizio", "data_fine"):
            valore = dt.fromisoformat(valore)
        elif mod.campo == "ore_stimate":
            valore = int(valore)

        ok = modifica_task(mod.task_id, **{mod.campo: valore})
        risultati.append({
            "task_id": mod.task_id,
            "campo": mod.campo,
            "applicato": ok,
        })

    # 2) Crea nuovi task
    nuovi_ids = []
    for nt in req.nuovi_task:
        new_id = aggiungi_task(
            progetto_id=req.progetto_id,
            nome=nt.nome,
            fase=nt.fase,
            ore_stimate=nt.ore_stimate,
            data_inizio=dt.fromisoformat(nt.data_inizio) if nt.data_inizio else get_oggi(),
            data_fine=dt.fromisoformat(nt.data_fine) if nt.data_fine else get_oggi(),
            stato=nt.stato,
            profilo_richiesto=nt.profilo_richiesto,
            dipendente_id=nt.dipendente_id,
            predecessore=nt.predecessore,
        )
        nuovi_ids.append(new_id)
        risultati.append({
            "task_id": new_id,
            "campo": "creato",
            "applicato": True,
        })

    # 3) Cambia stato progetto se richiesto
    stato_cambiato = False
    if req.cambia_stato_progetto and req.progetto_id:
        stato_cambiato = cambia_stato_progetto(req.progetto_id, req.cambia_stato_progetto)

    # 4) Ricalcola impatto post-applicazione
    impatto_post = calcola_impatto_saturazione(
        [{"task_id": r["task_id"], "campo": "stato", "nuovo_valore": "check"} for r in risultati if r["applicato"]],
        None
    )

    return {
        "risultati": risultati,
        "nuovi_task_ids": nuovi_ids,
        "stato_progetto_cambiato": stato_cambiato,
        "impatto_post": impatto_post,
    }


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: PIPELINE — SALVA BOZZA PIANIFICAZIONE
# ══════════════════════════════════════════════════════════════════════

class SalvaBozzaRequest(BaseModel):
    progetto_id: str
    dati_json: dict  # snapshot completo della tabella task in pianificazione


# Store bozze in memoria (in futuro: database)
BOZZE_STORE = {}


@app.post("/api/pianificazione/salva-bozza")
def salva_bozza(req: SalvaBozzaRequest, _: Utente = Depends(require_manager)):
    """Salva o aggiorna una bozza di pianificazione."""
    if PERSISTENT_MODE:
        salva_bozza_pianificazione(req.progetto_id, req.dati_json)
    else:
        BOZZE_STORE[req.progetto_id] = {
            "progetto_id": req.progetto_id,
            "dati_json": req.dati_json,
            "timestamp": get_oggi().strftime("%Y-%m-%d %H:%M"),
        }
    return {"salvato": True, "progetto_id": req.progetto_id}


@app.get("/api/pianificazione/bozza/{progetto_id}")
def carica_bozza(progetto_id: str, _: Utente = Depends(require_manager)):
    """Carica una bozza di pianificazione salvata."""
    if PERSISTENT_MODE:
        dati = carica_bozza_pianificazione(progetto_id)
        return {"progetto_id": progetto_id, "dati_json": dati}
    if progetto_id in BOZZE_STORE:
        return BOZZE_STORE[progetto_id]
    return {"progetto_id": progetto_id, "dati_json": None}


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: SALVA CONSUNTIVO
# ══════════════════════════════════════════════════════════════════════

class SalvaConsuntivoRequest(BaseModel):
    dipendente_id: str
    ore_per_task: dict[str, float] = {}
    stati_per_task: dict[str, str] = {}
    giorni_sede: int = 3
    giorni_remoto: int = 2
    ore_assenza: float = 0
    tipo_assenza: str = ""
    nota_assenza: str = ""
    spese: list[dict] = []


@app.post("/api/consuntivi/salva")
def salva_consuntivo_endpoint(req: SalvaConsuntivoRequest):
    """Salva il consuntivo settimanale di un dipendente."""
    try:
        dip = get_dipendente(req.dipendente_id)
    except (IndexError, KeyError):
        raise HTTPException(404, "Dipendente non trovato")

    if PERSISTENT_MODE:
        ok = salva_consuntivo(
            dipendente_id=req.dipendente_id,
            settimana=datetime.now(),
            ore_per_task=req.ore_per_task,
            stati_per_task=req.stati_per_task,
            giorni_sede=req.giorni_sede,
            giorni_remoto=req.giorni_remoto,
            ore_assenza=req.ore_assenza,
            tipo_assenza=req.tipo_assenza,
            nota_assenza=req.nota_assenza,
            spese_lista=req.spese if req.spese else None,
        )
        return {"salvato": ok, "dipendente": dip["nome"]}
    else:
        # Fallback: non salva ma conferma
        return {"salvato": True, "dipendente": dip["nome"], "nota": "Dati non persistenti (db non attivo)"}


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: IA SUGGERISCI TASK
# ══════════════════════════════════════════════════════════════════════

class SuggerisciTaskRequest(BaseModel):
    progetto_nome: str = ""
    progetto_cliente: str = ""
    descrizione: str = ""
    budget_ore: int = 0
    data_inizio: str = ""
    data_fine: str = ""


@app.post("/api/agent/suggerisci-task")
def suggerisci_task(req: SuggerisciTaskRequest):
    """Chiede all'agente di suggerire task per un progetto."""
    import json as json_mod

    contesto = {
        "progetto": req.progetto_nome,
        "cliente": req.progetto_cliente,
        "descrizione": req.descrizione,
        "budget_ore": req.budget_ore,
        "periodo": f"{req.data_inizio} — {req.data_fine}" if req.data_inizio else "Non specificato",
    }

    model = init_gemini(prompt_file="suggerisci_task.md")
    if model is None:
        raise HTTPException(503, "Agente AI non disponibile")

    contesto_json = json_mod.dumps(contesto, ensure_ascii=False, indent=2)
    prompt = f"Suggerisci i task per questo progetto:\n\n```json\n{contesto_json}\n```"

    try:
        chat = model.start_chat(history=[])
        response = chat.send_message(prompt)
        risposta_raw = response.text

        cleaned = risposta_raw.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()

        try:
            risultato = json_mod.loads(cleaned)
        except json_mod.JSONDecodeError:
            risultato = {"raw_response": risposta_raw, "parse_error": True}

        return risultato

    except Exception as e:
        raise HTTPException(500, f"Errore agente: {str(e)}")


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: IA VERIFICA PIANIFICAZIONE
# ══════════════════════════════════════════════════════════════════════

class VerificaPianificazioneRequest(BaseModel):
    progetto_nome: str = ""
    progetto_cliente: str = ""
    budget_ore: int = 0
    data_inizio: str = ""
    data_fine: str = ""
    task_pianificati: list[dict] = []


@app.post("/api/agent/verifica-pianificazione")
def verifica_pianificazione(req: VerificaPianificazioneRequest):
    """Chiede all'agente di verificare la pianificazione GANTT."""
    import json as json_mod

    # Costruisci contesto per l'agente
    contesto = {
        "progetto": {
            "nome": req.progetto_nome,
            "cliente": req.progetto_cliente,
            "budget_ore": req.budget_ore,
            "data_inizio": req.data_inizio,
            "data_fine": req.data_fine,
        },
        "task_pianificati": req.task_pianificati,
        "dipendenti_coinvolti": [],
    }

    # Aggiungi info saturazione per i dipendenti coinvolti
    nomi_coinvolti = set()
    for t in req.task_pianificati:
        if t.get("assegnato"):
            nomi_coinvolti.add(t["assegnato"])

    for nome in nomi_coinvolti:
        dip_match = _DIPENDENTI()[_DIPENDENTI()["nome"] == nome]
        if len(dip_match) > 0:
            d = dip_match.iloc[0]
            carico = carico_settimanale_dipendente(d["id"], get_oggi())
            contesto["dipendenti_coinvolti"].append({
                "nome": nome,
                "profilo": d["profilo"],
                "saturazione_attuale": round(carico / d["ore_sett"] * 100),
            })

    # Chiama Gemini
    model = init_gemini(prompt_file="verifica_pianificazione.md")
    if model is None:
        raise HTTPException(503, "Agente AI non disponibile")

    contesto_json = json_mod.dumps(contesto, ensure_ascii=False, indent=2)
    prompt = f"Verifica questa pianificazione GANTT:\n\n```json\n{contesto_json}\n```"

    try:
        chat = model.start_chat(history=[])
        response = chat.send_message(prompt)
        risposta_raw = response.text

        # Parsa JSON
        cleaned = risposta_raw.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()

        try:
            risultato = json_mod.loads(cleaned)
        except json_mod.JSONDecodeError:
            risultato = {"raw_response": risposta_raw, "parse_error": True}

        return risultato

    except Exception as e:
        raise HTTPException(500, f"Errore agente: {str(e)}")


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: EXPORT GANTT PDF
# ══════════════════════════════════════════════════════════════════════

from fastapi.responses import Response

@app.get("/api/gantt/export-pdf")
def export_gantt_pdf(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Genera e scarica un PDF del GANTT."""
    from gantt_pdf import genera_gantt_pdf

    # Riusa la logica di dati_gantt
    tasks = _TASKS().copy()
    if progetto_id:
        tasks = tasks[tasks["progetto_id"] == progetto_id]

    # Filtra solo progetti attivi (non sospesi)
    progetti_attivi = _PROGETTI()[~_PROGETTI()["stato"].isin(["Sospeso"])]["id"].tolist()
    if not progetto_id:
        tasks = tasks[tasks["progetto_id"].isin(progetti_attivi)]

    gantt_data = []
    for _, t in tasks.iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]]
        proj_nome = proj.iloc[0]["nome"] if len(proj) > 0 else "?"
        gantt_data.append({
            "id": t["id"],
            "name": t["nome"],
            "start": t["data_inizio"].strftime("%Y-%m-%d"),
            "end": t["data_fine"].strftime("%Y-%m-%d"),
            "project": proj_nome,
            "assignee": dip["nome"],
            "status": t["stato"],
            "estimated_hours": int(t["ore_stimate"]),
        })

    # Titolo
    if progetto_id:
        proj = _PROGETTI()[_PROGETTI()["id"] == progetto_id]
        titolo = f"GANTT — {proj.iloc[0]['nome']}" if len(proj) > 0 else "GANTT"
    else:
        titolo = "GANTT IMC-Group — Tutti i progetti"

    pdf_bytes = genera_gantt_pdf(gantt_data, titolo=titolo, progetto_filtro=progetto_id)

    filename = f"gantt_{progetto_id or 'tutti'}_{datetime.now().strftime('%Y%m%d')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/api/gantt/export-png")
def export_gantt_png(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Genera e scarica un PNG del GANTT (convertendo il PDF)."""
    from gantt_pdf import genera_gantt_pdf
    import subprocess
    import tempfile
    import os

    # Genera il PDF
    tasks = _TASKS().copy()
    if progetto_id:
        tasks = tasks[tasks["progetto_id"] == progetto_id]
    progetti_attivi = _PROGETTI()[~_PROGETTI()["stato"].isin(["Sospeso"])]["id"].tolist()
    if not progetto_id:
        tasks = tasks[tasks["progetto_id"].isin(progetti_attivi)]

    gantt_data = []
    for _, t in tasks.iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]]
        proj_nome = proj.iloc[0]["nome"] if len(proj) > 0 else "?"
        gantt_data.append({
            "id": t["id"], "name": t["nome"],
            "start": t["data_inizio"].strftime("%Y-%m-%d"),
            "end": t["data_fine"].strftime("%Y-%m-%d"),
            "project": proj_nome, "assignee": dip["nome"],
            "status": t["stato"], "estimated_hours": int(t["ore_stimate"]),
        })

    if progetto_id:
        proj = _PROGETTI()[_PROGETTI()["id"] == progetto_id]
        titolo = f"GANTT — {proj.iloc[0]['nome']}" if len(proj) > 0 else "GANTT"
    else:
        titolo = "GANTT IMC-Group — Tutti i progetti"

    pdf_bytes = genera_gantt_pdf(gantt_data, titolo=titolo)

    # Converti PDF → PNG con pdftoppm (poppler-utils)
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
            tmp_pdf.write(pdf_bytes)
            tmp_pdf_path = tmp_pdf.name

        png_path = tmp_pdf_path.replace(".pdf", "")
        subprocess.run(["pdftoppm", "-png", "-r", "200", "-singlefile", tmp_pdf_path, png_path],
                       check=True, capture_output=True)

        png_file = png_path + ".png"
        with open(png_file, "rb") as f:
            png_bytes = f.read()

        os.unlink(tmp_pdf_path)
        os.unlink(png_file)

        filename = f"gantt_{progetto_id or 'tutti'}_{datetime.now().strftime('%Y%m%d')}.png"
        return Response(content=png_bytes, media_type="image/png",
                        headers={"Content-Disposition": f"attachment; filename={filename}"})

    except (subprocess.CalledProcessError, FileNotFoundError):
        # Se pdftoppm non è disponibile, restituisci il PDF
        raise HTTPException(500, "Conversione PNG non disponibile. Installa poppler-utils: sudo apt install poppler-utils")


@app.get("/api/gantt/export-excel")
def export_gantt_excel(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Genera e scarica un file Excel con i dati GANTT + foglio visivo."""
    import io
    import pandas as pd

    tasks = _TASKS().copy()
    tasks = tasks[tasks["stato"] != "Eliminato"]
    if progetto_id:
        tasks = tasks[tasks["progetto_id"] == progetto_id]
    progetti_attivi = _PROGETTI()[~_PROGETTI()["stato"].isin(["Sospeso"])]["id"].tolist()
    if not progetto_id:
        tasks = tasks[tasks["progetto_id"].isin(progetti_attivi)]

    # Costruisci DataFrame per l'export
    export_data = []
    for _, t in tasks.iterrows():
        dip = get_dipendente(t["dipendente_id"])
        proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]]
        proj_nome = proj.iloc[0]["nome"] if len(proj) > 0 else "?"
        export_data.append({
            "Progetto": proj_nome,
            "Task": t["nome"],
            "Fase": t["fase"],
            "Assegnato a": dip["nome"],
            "Profilo": dip["profilo"],
            "Ore stimate": int(t["ore_stimate"]),
            "Data inizio": t["data_inizio"] if hasattr(t["data_inizio"], "strftime") else t["data_inizio"],
            "Data fine": t["data_fine"] if hasattr(t["data_fine"], "strftime") else t["data_fine"],
            "Stato": t["stato"],
        })

    df = pd.DataFrame(export_data)
    df = df.sort_values(["Progetto", "Data inizio"])

    buffer = io.BytesIO()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ═══ FOGLIO 1: Tabella dati ═══
        ws_data = wb.active
        ws_data.title = "Dati GANTT"

        # Header
        headers = ["Progetto", "Task", "Fase", "Assegnato a", "Profilo", "Ore stimate", "Data inizio", "Data fine", "Stato"]
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        for col, h in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Dati
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, h in enumerate(headers, 1):
                val = row[h]
                if h in ("Data inizio", "Data fine") and hasattr(val, "strftime"):
                    val = val.strftime("%d/%m/%Y")
                cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="left")

                # Colore riga per stato
                if row["Stato"] == "Completato":
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                elif row["Stato"] == "Da iniziare":
                    cell.fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")

        # Auto-width
        for col in range(1, len(headers) + 1):
            max_len = max(len(str(ws_data.cell(row=r, column=col).value or "")) for r in range(1, ws_data.max_row + 1))
            ws_data.column_dimensions[get_column_letter(col)].width = min(35, max(10, max_len + 2))

        # ═══ FOGLIO 2: GANTT Visivo ═══
        ws_gantt = wb.create_sheet("GANTT Visivo")

        # Trova range date
        date_inizio = []
        date_fine = []
        for _, t in tasks.iterrows():
            di = t["data_inizio"]
            df_t = t["data_fine"]
            if hasattr(di, "date"):
                date_inizio.append(di)
                date_fine.append(df_t)

        if not date_inizio:
            # Nessun task, foglio vuoto
            ws_gantt.cell(row=1, column=1, value="Nessun task da visualizzare")
        else:
            from datetime import timedelta as td
            min_date = min(date_inizio)
            max_date = max(date_fine)

            # Genera settimane
            settimane = []
            current = min_date - td(days=min_date.weekday())  # lunedì
            while current <= max_date + td(days=7):
                settimane.append(current)
                current += td(days=7)

            n_sett = len(settimane)

            # Colori per progetto
            colori_progetto = {
                "Adeguamento DORA": "4472C4",
                "Framework Compliance 262": "ED7D31",
                "Digitalizzazione Corpo Normativo": "70AD47",
                "Risk Assessment Operativo": "FFC000",
                "ProcessBook Aziendale": "9B59B6",
                "Attività Interne": "95A5A6",
            }
            colore_default = "5B9BD5"

            # Colori per stato
            colori_stato = {
                "Completato": "27ae60",
                "In corso": "3498db",
                "Da iniziare": "95a5a6",
                "Sospeso": "e67e22",
            }

            # Header: Task | Risorsa | Settimane...
            ws_gantt.cell(row=1, column=1, value="Task").font = Font(bold=True, size=9)
            ws_gantt.cell(row=1, column=2, value="Risorsa").font = Font(bold=True, size=9)
            ws_gantt.cell(row=1, column=3, value="Progetto").font = Font(bold=True, size=9)
            ws_gantt.cell(row=1, column=4, value="Stato").font = Font(bold=True, size=9)

            for i, sett in enumerate(settimane):
                cell = ws_gantt.cell(row=1, column=5 + i, value=sett.strftime("%d/%m"))
                cell.font = Font(size=7, bold=True)
                cell.alignment = Alignment(horizontal="center", text_rotation=90)
                ws_gantt.column_dimensions[get_column_letter(5 + i)].width = 4

            ws_gantt.column_dimensions["A"].width = 30
            ws_gantt.column_dimensions["B"].width = 18
            ws_gantt.column_dimensions["C"].width = 22
            ws_gantt.column_dimensions["D"].width = 12

            # Header fill
            for col in range(1, 5 + n_sett):
                ws_gantt.cell(row=1, column=col).fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
                ws_gantt.cell(row=1, column=col).font = Font(color="FFFFFF", bold=True, size=8)

            # Righe task
            row = 2
            current_project = ""
            for _, t in tasks.sort_values(["progetto_id", "data_inizio"]).iterrows():
                dip = get_dipendente(t["dipendente_id"])
                proj = _PROGETTI()[_PROGETTI()["id"] == t["progetto_id"]]
                proj_nome = proj.iloc[0]["nome"] if len(proj) > 0 else "?"

                # Riga separatore progetto
                if proj_nome != current_project:
                    current_project = proj_nome
                    sep_cell = ws_gantt.cell(row=row, column=1, value=proj_nome.upper())
                    sep_cell.font = Font(bold=True, size=9, color="FFFFFF")
                    proj_color = colori_progetto.get(proj_nome, colore_default)
                    for col in range(1, 5 + n_sett):
                        ws_gantt.cell(row=row, column=col).fill = PatternFill(start_color=proj_color, end_color=proj_color, fill_type="solid")
                    row += 1

                # Task info
                ws_gantt.cell(row=row, column=1, value=t["nome"]).font = Font(size=9)
                ws_gantt.cell(row=row, column=2, value=dip["nome"]).font = Font(size=8, color="666666")
                ws_gantt.cell(row=row, column=3, value=proj_nome).font = Font(size=8, color="666666")
                ws_gantt.cell(row=row, column=4, value=t["stato"]).font = Font(size=8)

                # Colore stato nella cella
                stato_color = colori_stato.get(t["stato"], "95a5a6")
                ws_gantt.cell(row=row, column=4).fill = PatternFill(start_color=stato_color, end_color=stato_color, fill_type="solid")
                ws_gantt.cell(row=row, column=4).font = Font(size=8, color="FFFFFF")

                # Barre GANTT
                task_start = t["data_inizio"]
                task_end = t["data_fine"]
                bar_color = colori_stato.get(t["stato"], colore_default)

                for i, sett in enumerate(settimane):
                    sett_end = sett + td(days=6)
                    if task_start <= sett_end and task_end >= sett:
                        ws_gantt.cell(row=row, column=5 + i).fill = PatternFill(
                            start_color=bar_color, end_color=bar_color, fill_type="solid"
                        )

                row += 1

            # Legenda in fondo
            row += 2
            ws_gantt.cell(row=row, column=1, value="Legenda:").font = Font(bold=True, size=9)
            row += 1
            for stato, colore in colori_stato.items():
                ws_gantt.cell(row=row, column=1, value=stato).font = Font(size=9)
                ws_gantt.cell(row=row, column=2).fill = PatternFill(start_color=colore, end_color=colore, fill_type="solid")
                row += 1

        # Salva
        wb.save(buffer)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    except ImportError:
        # Fallback CSV
        buffer = io.BytesIO()
        df_export = df.copy()
        df_export["Data inizio"] = df_export["Data inizio"].apply(lambda x: x.strftime("%d/%m/%Y") if hasattr(x, "strftime") else str(x))
        df_export["Data fine"] = df_export["Data fine"].apply(lambda x: x.strftime("%d/%m/%Y") if hasattr(x, "strftime") else str(x))
        df_export.to_csv(buffer, index=False, sep=";", encoding="utf-8-sig")
        media_type = "text/csv"
        ext = "csv"

    buffer.seek(0)
    filename = f"gantt_{progetto_id or 'tutti'}_{datetime.now().strftime('%Y%m%d')}.{ext}"
    return Response(content=buffer.getvalue(), media_type=media_type,
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


# ══════════════════════════════════════════════════════════════════════
# ENDPOINT: SCENARIO — TAVOLO DI LAVORO
# ══════════════════════════════════════════════════════════════════════
 
class ModificaScenario(BaseModel):
    """Singola modifica nello scenario."""
    tipo: str                          # "sposta_task" | "cambia_focus"
    # Per sposta_task:
    task_id: str = ""
    nuovo_inizio: str = ""             # ISO date
    nuova_fine: str = ""               # ISO date
    nuove_ore: int = 0
    # Per cambia_focus:
    dipendente_id: str = ""
    progetto_focus: str = ""
    percentuale: int = 100
    durata_settimane: int = 2
    data_inizio_focus: str = ""
 
 
class SimulaRequest(BaseModel):
    """Richiesta di simulazione scenario."""
    modifiche: list[ModificaScenario]
 
 
class ConfermaRequest(BaseModel):
    """Richiesta di conferma e applicazione scenario."""
    modifiche: list[ModificaScenario]
 
 
class InterpretaRequest(BaseModel):
    """Richiesta di interpretazione linguaggio naturale."""
    testo: str
    contesto_extra: str = ""
 
 
# ── POST /api/scenario/simula ─────────────────────────────────────────
 
@app.post("/api/scenario/simula")
def scenario_simula(req: SimulaRequest, _: Utente = Depends(require_manager)):
    """
    Simula uno scenario SENZA modificare il database.
    Riceve modifiche, calcola cascata, restituisce GANTT prima/dopo
    + conseguenze + saturazioni.
    """
    from scenario_engine import simula_scenario, _to_date
 
    # Converti request → formato motore
    modifiche = []
    for mod in req.modifiche:
        m = {"tipo": mod.tipo}
        if mod.tipo == "sposta_task":
            m["task_id"] = mod.task_id
            if mod.nuovo_inizio:
                m["nuovo_inizio"] = mod.nuovo_inizio
            if mod.nuova_fine:
                m["nuova_fine"] = mod.nuova_fine
            if mod.nuove_ore > 0:
                m["nuove_ore"] = mod.nuove_ore
        elif mod.tipo == "cambia_focus":
            m["dipendente_id"] = mod.dipendente_id
            m["progetto_focus"] = mod.progetto_focus
            m["percentuale"] = mod.percentuale
            m["durata_settimane"] = mod.durata_settimane
            if mod.data_inizio_focus:
                m["data_inizio_focus"] = mod.data_inizio_focus
        modifiche.append(m)
 
    # Esegui simulazione
    risultato = simula_scenario(
        _TASKS(), _DIPENDENTI(), _PROGETTI(),
        modifiche, data_oggi=get_oggi()
    )
 
    # Costruisci GANTT prima/dopo per i progetti impattati
    gantt_prima = {}
    gantt_dopo = {}
    progetti_impattati = set()
 
    for tid in risultato["task_modificati"]:
        if tid in risultato["tasks_dopo"]:
            pid = risultato["tasks_dopo"][tid].get("progetto_id", "")
            if pid:
                progetti_impattati.add(pid)
 
    for pid in progetti_impattati:
        proj = get_progetto(pid)
 
        # GANTT prima
        tasks_prima_prog = []
        for t in risultato["tasks_prima"].values():
            if t.get("progetto_id") != pid:
                continue
            t_inizio = _to_date(t.get("data_inizio"))
            t_fine = _to_date(t.get("data_fine"))
            tasks_prima_prog.append({
                "id": t["id"],
                "name": t.get("nome", ""),
                "start": t_inizio.isoformat() if t_inizio else "",
                "end": t_fine.isoformat() if t_fine else "",
                "status": t.get("stato", ""),
                "assignee": get_dipendente(t.get("dipendente_id", ""))["nome"],
                "estimated_hours": t.get("ore_stimate", 0),
                "project": proj["nome"],
            })
 
        # GANTT dopo
        tasks_dopo_prog = []
        for t in risultato["tasks_dopo"].values():
            if t.get("progetto_id") != pid:
                continue
            t_inizio = _to_date(t.get("data_inizio"))
            t_fine = _to_date(t.get("data_fine"))
            tasks_dopo_prog.append({
                "id": t["id"],
                "name": t.get("nome", ""),
                "start": t_inizio.isoformat() if t_inizio else "",
                "end": t_fine.isoformat() if t_fine else "",
                "status": t.get("stato", ""),
                "assignee": get_dipendente(t.get("dipendente_id", ""))["nome"],
                "estimated_hours": t.get("ore_stimate", 0),
                "project": proj["nome"],
            })
 
        gantt_prima[pid] = {
            "progetto": proj["nome"],
            "cliente": proj.get("cliente", ""),
            "tasks": tasks_prima_prog,
        }
        gantt_dopo[pid] = {
            "progetto": proj["nome"],
            "cliente": proj.get("cliente", ""),
            "tasks": tasks_dopo_prog,
        }
 
    return {
        "gantt_prima": gantt_prima,
        "gantt_dopo": gantt_dopo,
        "conseguenze": risultato["conseguenze"],
        "scadenze_bucate": risultato["scadenze_bucate"],
        "saturazioni": risultato["saturazioni"],
        "n_task_modificati": len(risultato["task_modificati"]),
        "progetti_impattati": [
            {"id": pid, "nome": get_progetto(pid)["nome"]}
            for pid in progetti_impattati
        ],
    }
 
 
# ── POST /api/scenario/conferma ───────────────────────────────────────
 
@app.post("/api/scenario/conferma")
def scenario_conferma(req: ConfermaRequest, _: Utente = Depends(require_manager)):
    """
    Applica le modifiche dello scenario al database.
    Prima ri-simula per ottenere le date propagate,
    poi scrive tutto nel db.
    """
    from scenario_engine import simula_scenario, _to_date
    from datetime import datetime as dt
 
    # Converti request → formato motore (stessa logica di /simula)
    modifiche = []
    for mod in req.modifiche:
        m = {"tipo": mod.tipo}
        if mod.tipo == "sposta_task":
            m["task_id"] = mod.task_id
            if mod.nuovo_inizio:
                m["nuovo_inizio"] = mod.nuovo_inizio
            if mod.nuova_fine:
                m["nuova_fine"] = mod.nuova_fine
            if mod.nuove_ore > 0:
                m["nuove_ore"] = mod.nuove_ore
        elif mod.tipo == "cambia_focus":
            m["dipendente_id"] = mod.dipendente_id
            m["progetto_focus"] = mod.progetto_focus
            m["percentuale"] = mod.percentuale
            m["durata_settimane"] = mod.durata_settimane
            if mod.data_inizio_focus:
                m["data_inizio_focus"] = mod.data_inizio_focus
        modifiche.append(m)
 
    risultato = simula_scenario(
        _TASKS(), _DIPENDENTI(), _PROGETTI(),
        modifiche, data_oggi=get_oggi()
    )
 
    # Applica al db ogni task che è cambiato
    applicati = []
    errori = []
 
    for tid in risultato["task_modificati"]:
        task_dopo = risultato["tasks_dopo"].get(tid)
        task_prima = risultato["tasks_prima"].get(tid)
        if not task_dopo or not task_prima:
            continue
 
        kwargs = {}
        nuovo_inizio = _to_date(task_dopo.get("data_inizio"))
        vecchio_inizio = _to_date(task_prima.get("data_inizio"))
        nuova_fine = _to_date(task_dopo.get("data_fine"))
        vecchia_fine = _to_date(task_prima.get("data_fine"))
 
        if nuovo_inizio and vecchio_inizio and nuovo_inizio != vecchio_inizio:
            kwargs["data_inizio"] = dt.combine(nuovo_inizio, dt.min.time())
        if nuova_fine and vecchia_fine and nuova_fine != vecchia_fine:
            kwargs["data_fine"] = dt.combine(nuova_fine, dt.min.time())
 
        nuove_ore = task_dopo.get("ore_stimate")
        vecchie_ore = task_prima.get("ore_stimate")
        if nuove_ore and vecchie_ore and nuove_ore != vecchie_ore:
            kwargs["ore_stimate"] = nuove_ore
 
        if kwargs:
            ok = modifica_task(tid, **kwargs)
            if ok:
                applicati.append({
                    "task_id": tid,
                    "task_nome": task_dopo.get("nome", ""),
                    "progetto": get_progetto(task_dopo.get("progetto_id", ""))["nome"],
                    "modifiche_applicate": {k: str(v) for k, v in kwargs.items()},
                })
            else:
                errori.append({
                    "task_id": tid,
                    "task_nome": task_dopo.get("nome", ""),
                    "errore": "Modifica non riuscita",
                })
 
    return {
        "successo": len(errori) == 0,
        "n_applicati": len(applicati),
        "applicati": applicati,
        "errori": errori,
        "conseguenze_applicate": risultato["conseguenze"],
    }
 
 
# ── POST /api/scenario/interpreta ─────────────────────────────────────
 
@app.post("/api/scenario/interpreta")
def scenario_interpreta(req: InterpretaRequest, _: Utente = Depends(require_manager)):
    """
    L'IA interpreta il linguaggio naturale del management
    e lo traduce in modifiche strutturate per /scenario/simula.
    NON esegue la simulazione — restituisce le modifiche proposte
    che il frontend mostrerà al management per conferma/modifica.
    """
    import json as json_mod
 
    model = init_gemini(prompt_file="interpreta_scenario.md")
    if model is None:
        raise HTTPException(503, "Agente AI non disponibile")
 
    contesto = get_contesto_ia()
    contesto_json = json_mod.dumps(contesto, ensure_ascii=False, indent=2)
 
    prompt = f"Il management dice:\n\n\"{req.testo}\"\n\n"
    if req.contesto_extra:
        prompt += f"Contesto aggiuntivo dal management: {req.contesto_extra}\n\n"
    prompt += f"Stato attuale del sistema:\n\n```json\n{contesto_json}\n```"
 
    try:
        chat = model.start_chat(history=[])
        response = chat.send_message(prompt)
        risposta_raw = response.text
 
        # Pulisci e parsa JSON
        cleaned = risposta_raw.strip()
        if cleaned.startswith("```json"):
            cleaned = cleaned[7:]
        if cleaned.startswith("```"):
            cleaned = cleaned[3:]
        if cleaned.endswith("```"):
            cleaned = cleaned[:-3]
        cleaned = cleaned.strip()
 
        try:
            risultato = json_mod.loads(cleaned)
        except json_mod.JSONDecodeError:
            risultato = {
                "interpretazione": risposta_raw,
                "modifiche": [],
                "domande": "",
                "note_contesto": "",
                "parse_error": True,
            }
 
        return risultato
 
    except Exception as e:
        raise HTTPException(500, f"Errore agente: {str(e)}")

# ══════════════════════════════════════════════════════════════════
# ENDPOINT: ATTIVITÀ INTERNE (P010)
# ══════════════════════════════════════════════════════════════════

@app.post("/api/attivita-interne")
def crea_attivita_interna(req: AttivitaInternaRequest):
    """Crea un task su P010 (Attività Interne) per un dipendente."""
    try:
        dip = get_dipendente(req.dipendente_id)
    except (IndexError, KeyError):
        raise HTTPException(404, "Dipendente non trovato")
 
    p010 = _PROGETTI()[_PROGETTI()["id"] == "P010"]
    if p010.empty:
        raise HTTPException(400, "Progetto P010 non trovato")
 
    from datetime import datetime as dt
    import time
    time.sleep(0.3)
 
    new_id = aggiungi_task(
        progetto_id="P010",
        nome=req.nome,
        fase=req.categoria,
        ore_stimate=req.ore_stimate if req.ore_stimate > 0 else req.ore_settimanali * 20,
        data_inizio=dt.fromisoformat(req.data_inizio),
        data_fine=dt.fromisoformat(req.data_fine),
        stato="In corso",
        profilo_richiesto=dip.get("profilo", ""),
        dipendente_id=req.dipendente_id,
    )
 
    return {"ok": True, "task_id": new_id, "messaggio": f"Attività '{req.nome}' creata per {dip['nome']}"}
 
 
@app.delete("/api/attivita-interne/{task_id}")
def elimina_attivita_interna(task_id: str):
    """Elimina (soft) un task di attività interna (solo P010)."""
    tasks = _TASKS()
    task = tasks[tasks["id"] == task_id]
    if task.empty:
        raise HTTPException(404, "Task non trovato")
    if task.iloc[0]["progetto_id"] != "P010":
        raise HTTPException(400, "Solo task di Attività Interne possono essere eliminati da qui")

    ok = modifica_task(task_id, stato="Eliminato")
    if ok:
        return {"ok": True, "messaggio": f"Task {task_id} eliminato"}
    raise HTTPException(500, "Errore nell'eliminazione")


@app.patch("/api/tasks/{task_id}/elimina")
def elimina_task_generico(task_id: str, _: Utente = Depends(require_manager)):
    """Elimina (soft) qualsiasi task cambiando lo stato a 'Eliminato'."""
    tasks = _TASKS()
    task = tasks[tasks["id"] == task_id]
    if task.empty:
        raise HTTPException(404, "Task non trovato")

    task_nome = task.iloc[0]["nome"]
    ok = modifica_task(task_id, stato="Eliminato")
    if ok:
        return {"ok": True, "messaggio": f"Task '{task_nome}' eliminato"}
    raise HTTPException(500, "Errore nell'eliminazione")


# ══════════════════════════════════════════════════════════════════════
# AVVIO
# ══════════════════════════════════════════════════════════════════════

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)