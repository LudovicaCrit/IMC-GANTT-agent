"""
═══════════════════════════════════════════════════════════════════════════
backend/routes/gantt.py — Router per endpoint /api/gantt
═══════════════════════════════════════════════════════════════════════════

SCOPO
─────
Espone gli endpoint per la pagina GANTT del frontend:
  - Lettura dati GANTT in JSON (per rendering frappe-gantt nel frontend)
  - Esportazione GANTT in formato PDF, PNG e Excel (download file)
Tutti gli endpoint sono manager-only (vista d'insieme aziendale).

Decisione di design (5 maggio 2026): tutti gli endpoint /api/gantt/*
in un unico file. Coerente con il principio "1 prefix = 1 file" e con la
discoverability per chi cerca dove vivono i GANTT. Se in futuro arriveranno
altri export non-GANTT (es. Excel di marginalità, PDF di consuntivi),
si valuterà di nuovo la scorporazione in un eventuale `routes/export.py`.

ENDPOINT ESPOSTI
────────────────
┌──────────────────────────────────┬──────────┬──────────────────────────────┐
│ Path                             │ Metodo   │ Auth                         │
├──────────────────────────────────┼──────────┼──────────────────────────────┤
│ /api/gantt                       │ GET      │ require_manager              │
│ /api/gantt/strutturato           │ GET      │ require_manager              │
│ /api/gantt/export-pdf            │ GET      │ require_manager              │
│ /api/gantt/export-png            │ GET      │ require_manager              │
│ /api/gantt/export-excel          │ GET      │ require_manager              │
└──────────────────────────────────┴──────────┴──────────────────────────────┘

DETTAGLIO ENDPOINT
──────────────────
1. GET /api/gantt?progetto_id=P00X
   - Manager-only.
   - Parametro opzionale: filtra per progetto.
   - Esclude task con stato "Eliminato" (soft delete).
   - Restituisce JSON formattato per la libreria frappe-gantt:
     id, name, start, end, progress (calcolato da ore_consuntivate/ore_stimate),
     dependencies (predecessore), project, project_id, assignee, profile,
     status, estimated_hours, hours_done, predecessor_name.

2. GET /api/gantt/export-pdf?progetto_id=P00X
   - Manager-only.
   - Genera un PDF tramite `gantt_pdf.genera_gantt_pdf` (modulo dedicato).
   - Filtra automaticamente progetti non sospesi se non c'è filtro esplicito.
   - Response: application/pdf con Content-Disposition: attachment.

3. GET /api/gantt/export-png?progetto_id=P00X
   - Manager-only.
   - Genera prima il PDF, poi lo converte in PNG via `pdftoppm` (poppler-utils).
   - Risoluzione: 200 dpi, single page.
   - 500 con messaggio specifico se `pdftoppm` non è installato.

4. GET /api/gantt/export-excel?progetto_id=P00X
   - Manager-only.
   - Genera un .xlsx con due fogli:
     • "Dati GANTT": tabella con stati colorati per riga
     • "GANTT Visivo": rappresentazione a barre settimanali, raggruppate
       per progetto, con legenda stati
   - Fallback CSV se openpyxl non è installato.

PATTERN AUTH USATI
──────────────────
- `require_manager` su tutti e 4 gli endpoint. Il GANTT aziendale e i suoi
  export sono informazione manageriale (Scenario B); Helena vede solo i
  propri task tramite `/api/tasks` con filtro Scenario B.

DIPENDENZE ESTERNE (oltre al backend)
─────────────────────────────────────
- `gantt_pdf` (modulo locale): `genera_gantt_pdf()` — usato da PDF e PNG
- `reportlab` (via gantt_pdf): rendering PDF
- `pdftoppm` (binario poppler-utils): conversione PDF→PNG. Sistema Linux:
  `sudo apt install poppler-utils`. Se manca, l'endpoint risponde 500 con
  messaggio chiaro.
- `openpyxl`: rendering Excel. Se manca, fallback su CSV.
- `pandas`: usato come strumento di sort/iter dentro l'export Excel
  (NON più come cache di lettura — i dati arrivano da Postgres).

📌 TODO Blocco 2 roadmap (Macchina delle Fasi):
   Riprogettare `dati_gantt` per restituire dati strutturati a livello di
   fase, non solo task piatti. Il GANTT diventerà "barre raggruppate per
   fase, espandibili al dettaglio task" — coerente con la vista per fasi
   discussa con Vincenzo (post-Francesco).
   Gli export andranno adattati di conseguenza.

DIPENDENZE INTERNE
──────────────────
- `data` (modulo): `get_dipendente`.
- `models`: `Progetto`, `Fase`, `Task`, `Consuntivo`, `Dipendente`,
  `get_session`, `STATI_PROGETTO_ATTIVI`.
- `data_db_impl._to_dt`: normalizza `Date` SQL → `datetime` a mezzanotte
  (necessario per l'export Excel — vedi NOTE TECNICHE).
- `deps`: `require_manager`.
- `models.Utente` per type hint.

NOTE TECNICHE
─────────────
**Date e formato datetime per l'export Excel.** Il foglio "GANTT Visivo"
fa `if hasattr(di, "date")` per validare i task con date valide. Un
oggetto `date` puro NON ha l'attributo `.date` (ce l'ha `datetime`),
mentre lo storico pandas.Timestamp ce l'aveva. Inoltre i confronti
`task_start <= sett_end` (datetime vs datetime nel codice originale)
solleverebbero TypeError se mescolassimo date e datetime. Per iso-
comportamento, le date dei task sono **pre-convertite a datetime via
_to_dt()** quando costruiamo `task_records`.

STORIA
──────
Estratto da main.py il 5 maggio 2026 nell'ambito del refactoring strangler.
Tutti e 4 gli endpoint /api/gantt/* sono qui (decisione presa con Ludovica
durante il refactoring stesso).
Letture DataFrame migrate a Postgres diretto il 21 maggio 2026 (handoff
migrazione §6-ter). `gantt_strutturato` era già nativamente su SQLAlchemy.
═══════════════════════════════════════════════════════════════════════════
"""

from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Response
from sqlalchemy.orm import joinedload, selectinload
from sqlalchemy import func

from deps import require_manager
from models import (
    Utente, Progetto, Fase, Task, Consuntivo, Dipendente, Sottotask,
    get_session, STATI_PROGETTO_ATTIVI, urgenza_fase_risolta,
)


# ─────────────────────────────────────────────────────────────────────────
# Step 3.1 (Gruppo A, 28/05/2026): la colonna `Task.predecessore` non esiste
# più — le dipendenze vivono nella tabella-grafo `dipendenza_task`
# (relationship `Task.dipendenze_entranti`). Finché il GANTT/agent espongono
# UN solo predecessore (iso-comportamento col vecchio campo stringa), questa
# helper restituisce l'id del predecessore "principale" del task: la prima
# dipendenza entrante di tipo FS, o in mancanza la prima entrante in assoluto.
# Restituisce "" se il task non ha predecessori.
#
# STRADA 2 (Code): quando il motore/payload passeranno a dipendenze MULTIPLE
# tipizzate, questa helper va sostituita dall'esposizione dell'intera lista
# `dipendenze_entranti` (id + tipo_dipendenza). Vedi HANDOFF Gruppo A→B.
# ─────────────────────────────────────────────────────────────────────────
def _semaforo_payload(nodo):
    """Nodo dell'albero di `semaforo_progetti` → sotto-oggetto del payload.

    Traduce la chiave `semaforo` del nodo in `colore`: annidando il nodo così
    com'è si otterrebbe `"semaforo": {"semaforo": "rosso", ...}`, che si legge
    male e si sbaglia a scrivere. Scarta anche le chiavi STRUTTURALI del nodo
    (`fasi`, `task`, `sottotask`): il payload ha già la propria gerarchia, e
    duplicarla dentro ogni semaforo la raddoppierebbe. Lo scarto è di FORMA, non
    di contenuto — il sotto-dict `sottotask` del nodo-task viene letto a parte
    dal loop, che lo appaia alle righe Sottotask vere e ne ricava il colore di
    ciascun pezzo (lavoro B): il colore dei sottotask non si ricalcola, si
    riprende da qui.

    `nodo` assente → None, come fa `scostamento_per_task.get(t.id)` per i task
    senza niente da dire. NON un KeyError: `semaforo_progetti` cammina la
    gerarchia per conto suo e applica gli stessi filtri di questo endpoint
    (task «Eliminato» esclusi di là come di qua), quindi ogni unità serializzata
    DEVE trovare il suo nodo — ma se un domani i due filtri divergessero,
    perdere un colore su una riga è meno grave che restituire 500 sull'intero
    drill-down di quattro pagine. Che la copertura sia totale non è affidato
    alla speranza: c'è un test che verifica che nessuna unità del payload esca
    con `semaforo: null`, e lì la divergenza sarebbe rumorosa.
    """
    if not nodo:
        return None
    return {
        "colore": nodo["semaforo"],
        "origine": nodo["origine"],
        "figli_rossi": nodo["figli_rossi"],
    }


def _predecessore_principale(task: Task) -> str:
    """Id del predecessore principale del task (Strada 1, singolo).

    Richiede che `task.dipendenze_entranti` sia caricata (selectinload).
    """
    entranti = task.dipendenze_entranti or []
    if not entranti:
        return ""
    fs = [d for d in entranti if d.tipo_dipendenza == "FS"]
    scelta = fs[0] if fs else entranti[0]
    return scelta.task_predecessore_id or ""
from data import get_dipendente, scostamento_stime_sottotask, semaforo_progetti
from data_db_impl import _to_dt


# ── Router ───────────────────────────────────────────────────────────────
router = APIRouter(prefix="/api/gantt", tags=["gantt"])


# ═════════════════════════════════════════════════════════════════════════
# 1. GET /api/gantt — dati strutturati per il rendering frontend
# ═════════════════════════════════════════════════════════════════════════

@router.get("")
def dati_gantt(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Restituisce i dati formattati per il componente GANTT del frontend."""
    session = get_session()
    q = session.query(Task).options(
        joinedload(Task.progetto),
        # Step 3.1 (Gruppo A): le dipendenze ora sono in dipendenza_task;
        # le carico qui per leggere il predecessore principale senza N+1.
        selectinload(Task.dipendenze_entranti),
    ).filter(Task.stato != "Eliminato")
    if progetto_id:
        q = q.filter(Task.progetto_id == progetto_id)
    tasks = q.all()

    # Ore consuntivate per task in UNA query (GROUP BY task_id)
    task_ids = [t.id for t in tasks]
    ore_per_task = {}
    if task_ids:
        rows = session.query(
            Consuntivo.task_id,
            func.coalesce(func.sum(Consuntivo.ore_dichiarate), 0)
        ).filter(Consuntivo.task_id.in_(task_ids)).group_by(Consuntivo.task_id).all()
        ore_per_task = {tid: float(ore) for tid, ore in rows}

    # Nomi predecessori in UNA query (sostituisce il lookup _TASKS()[...id == pred])
    # Step 3.1 (Gruppo A): il predecessore "principale" viene dalla tabella-grafo
    # via _predecessore_principale (prima dip. entrante FS), non più da t.predecessore.
    pred_per_task = {t.id: _predecessore_principale(t) for t in tasks}
    pred_ids = [p for p in pred_per_task.values() if p]
    nomi_pred = {}
    if pred_ids:
        pred_rows = session.query(Task.id, Task.nome).filter(Task.id.in_(pred_ids)).all()
        nomi_pred = {pid: pn for pid, pn in pred_rows}
    session.close()

    result = []
    for t in tasks:
        dip = get_dipendente(t.dipendente_id)
        ore_cons = ore_per_task.get(t.id, 0.0)
        ore_stimate = int(t.ore_stimate) if t.ore_stimate else 0  # display (stima storica, gruppo B)
        ore_pianificate = int(t.ore_pianificate) if t.ore_pianificate else 0  # piano corrente (#3)

        # Progress calcolato su ore consuntivate / ore PIANIFICATE (piano corrente).
        # Migrazione #3 passo 2 (#3). Post-backfill pianificate == stimate → oracolo
        # progress invariato.
        if t.stato == "Completato":
            progress = 100
        elif ore_pianificate > 0 and ore_cons > 0:
            progress = min(99, round(ore_cons / ore_pianificate * 100))
        else:
            progress = 0

        pred = pred_per_task.get(t.id, "")
        result.append({
            "id": t.id,
            "name": t.nome,
            # Date opzionali: un task senza data compare comunque nel GANTT
            # con stringa vuota (frappe-gantt lo ignora visivamente, ma il
            # task resta nel payload — niente 500).
            "start": t.data_inizio.strftime("%Y-%m-%d") if t.data_inizio else "",
            "end": t.data_fine.strftime("%Y-%m-%d") if t.data_fine else "",
            "progress": progress,
            "dependencies": pred,
            "predecessor_name": nomi_pred.get(pred, "") if pred else "",
            "project": t.progetto.nome if t.progetto else "",
            "project_id": t.progetto_id,
            "assignee": dip["nome"],
            "profile": dip["profilo"],
            "status": t.stato,
            "estimated_hours": ore_stimate,
            "hours_done": round(ore_cons, 1),
        })

    return result


# ═════════════════════════════════════════════════════════════════════════
# 1.b GET /api/gantt/strutturato — gerarchia Progetto → Fase → Task
# ═════════════════════════════════════════════════════════════════════════

@router.get("/strutturato")
def gantt_strutturato(
    stato: Optional[str] = None,
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Restituisce la gerarchia Progetto → Fase → Task per il drill-down GANTT.

    Step 2.2 del Blocco 2 esteso (handoff v15 §2.2 punto 1, §2.3).

    Filtri query params:
    - `stato`: filtra i progetti per stato. Default: "attivi"
      (In esecuzione + Sospeso, allineato a handoff §3.3).
      Valori: "attivi" | "all" | "bozza" | "in esecuzione" | ...
    - `progetto_id`: drill su un singolo progetto (utile per `/cantiere/{id}`).

    Struttura risposta:
    [
      {
        "id": "P001", "nome": "...", "cliente": "...", "stato": "...",
        "data_inizio": "...", "data_fine": "...",
        "ore_vendute_totali": 240, "ore_consumate_totali": 130,
        "semaforo": {"colore": "rosso", "origine": "figli", "figli_rossi": 1},
        "fasi": [
          {
            "id": 1, "nome": "Analisi", "ordine": 1, "stato": "In corso",
            "data_inizio": "...", "data_fine": "...",
            "ore_vendute": 80, "ore_consumate": 50,
            "semaforo": {"colore": "rosso", "origine": "figli", "figli_rossi": 2},
            "tasks": [
              {
                "id": "T001", "nome": "...", "stato": "...",
                "ore_stimate": 40, "ore_consumate": 25,
                "data_inizio": "...", "data_fine": "...",
                "dipendente_id": "...", "dipendente_nome": "...",
                "semaforo": {"colore": "rosso", "origine": "entrambe",
                             "figli_rossi": 1},
                "predecessore": "",
                "sottotask": [            # solo sui task scomposti
                  {"id": 7, "nome": "...", "ordine": 1, "stato": "Da iniziare",
                   "ore_stimate": 8, "dipendente_id": null,
                   "dipendente_nome": "",
                   "semaforo": {"colore": "rosso", "origine": "propria",
                                "figli_rossi": 0}}, ...
                ]
              }, ...
            ]
          }, ...
        ]
      }, ...
    ]

    Design:
    - L'endpoint restituisce dati RAW più due GIUDIZI CALCOLATI nello strato
      dati: `scostamento` (Step 2.3) e `semaforo` (semaforo ritardabilità,
      strato 1). Fino al 30/07/2026 questa riga diceva «l'endpoint è stupido,
      il frontend calcola progress %, colori e default aperture»: `scostamento`
      l'aveva già incrinata, il semaforo la rende falsa, e vale la pena dire
      perché invece di lasciarla lì.
      Il semaforo NON può stare nel frontend. Serve la GERARCHIA per aggregare
      il peggio-dei-figli lungo sottotask → task → fase → progetto, e in strato
      2 servirà la STORIA delle percentuali dichiarate — cioè query, non dati
      che il payload contiene. Calcolarlo di là significherebbe o mandare al
      client tutto quello che serve al calcolo, o riscriverlo in JavaScript e
      farlo divergere da quello del backend alla prima modifica.
      Resta invece vero, e non è cambiato, che il frontend decide la RESA:
      quali colori disegnare, quali fasi aprire per default (handoff §2.3
      «fasi In corso aperte» è scelta UI), come graduare un rosso proprio da un
      rosso ereditato. Il backend dice il GIUDIZIO, il frontend come si vede.
    - Performance: joinedload per evitare N+1 query su fasi e task.
    - Aggregazioni ore: una query sum() sui consuntivi per evitare
      iterazioni Python.
    - Semaforo: UNA chiamata batch per tutti i progetti in scope (2 query),
      non una per unità. Stesso pattern di `scostamento`.

    IL CAMPO `semaforo` (a progetto, fase e task):
        "semaforo": {"colore": "rosso"|"giallo"|"grigio"|"verde",
                     "origine": "propria"|"figli"|"entrambe"|null,
                     "figli_rossi": int}
    Sotto-oggetto e non tre campi piatti, per la stessa ragione di
    `scostamento`: è UN fatto con più facce, e in strato 2 gli si affiancheranno
    altre (le ore residue, il motivo del giallo) senza allargare di nuovo lo
    spazio dei nomi al primo livello.
    `origine` dice DA DOVE viene il colore — dal calendario di questo livello
    ("propria"), da un figlio ("figli"), o da entrambi. È ciò che permette al
    frontend di graduare la resa senza che il backend inventi soglie: un
    progetto rosso con 60 giorni di margine e un task scaduto dentro è
    "figli", e va disegnato diversamente da uno rosso di suo.
    `figli_rossi` conta i figli DIRETTI rossi (le fasi per il progetto, i task
    per la fase): risponde a «dove clicco adesso».
    I SOTTOTASK (lavoro B) — quarto livello, annidato nel task:
        "sottotask": [{id, nome, ordine, stato, ore_stimate,
                       dipendente_id, dipendente_nome, semaforo}, ...]
    La chiave c'è SOLO sui task scomposti: assente = «questo task non ha
    pezzi», ed è la convenzione della casa (`task_settimana_dipendente`).
    Ordinati per (ordine, id), come `lista_sottotask_task`.
    I pezzi «Annullato» non compaiono: sono stati tolti dal piano, e un GANTT
    disegna il piano. I «Sospeso» sì — in pausa, ma ancora nel piano.

    NIENTE DATE sui sottotask, e non è una dimenticanza: `Sottotask` non ne ha
    PER SCELTA (models.py: «eredita la finestra temporale del task padre»). La
    finestra di un pezzo è quella del task che lo racchiude, che il frontend ha
    già sotto mano nell'oggetto immediatamente superiore. Copiarle qui creerebbe
    una seconda verità destinata a divergere alla prima modifica delle date del
    task — e, se un domani i sottotask avessero date proprie, un payload che ne
    porta già di ereditate cambierebbe significato in silenzio.
    Il colore del pezzo, però, RIFLETTE quella data ereditata: un pezzo vivo
    dentro un task scaduto è rosso.

    NIENTE PERCENTUALE, per ora. Serve `_baseline_percentuali(tipo="sottotask")`
    — una query in più e una decisione su dove viva «avanzamento corrente» — e
    oggi uscirebbe `null` per ogni pezzo: in DB non c'è UNA riga con
    `percentuale` non-NULL, né in `consuntivi` né in `consuntivo_sottotask`.
    Si aggiunge col lavoro A, quando le barre andranno disegnate davvero.
    """
    session = get_session()
    try:
        # ── 1. Query progetti con filtro stato ────────────────────────
        q = session.query(Progetto).options(
            joinedload(Progetto.fasi).joinedload(Fase.task)
                .selectinload(Task.dipendenze_entranti)
        )
        if progetto_id:
            q = q.filter(Progetto.id == progetto_id)
        elif stato is None or stato.lower() == "attivi":
            q = q.filter(Progetto.stato.in_(STATI_PROGETTO_ATTIVI))
        elif stato.lower() != "all":
            q = q.filter(func.lower(Progetto.stato) == stato.lower())
        progetti = q.order_by(Progetto.id).all()

        # ── 2. Tutti i consuntivi in una query ────────────────────────
        # Aggrego per task_id per non fare N+1 nel loop sotto.
        # Filtriamo i task "Eliminato" (soft delete): non devono comparire
        # nel drill-down né contribuire ai conteggi (Step 2.4-bis fix).
        task_ids_all = [
            t.id for p in progetti for f in p.fasi for t in f.task
            if t.stato != "Eliminato"
        ]
        ore_per_task = {}
        if task_ids_all:
            righe = session.query(
                Consuntivo.task_id,
                func.coalesce(func.sum(Consuntivo.ore_dichiarate), 0)
            ).filter(Consuntivo.task_id.in_(task_ids_all)).group_by(Consuntivo.task_id).all()
            ore_per_task = {tid: float(ore) for tid, ore in righe}

        # ── 2-bis. Scostamento stime sottotask, in UNA chiamata ───────
        # Step 2.3 sottotask (30/07/2026). Stessa logica di ore_per_task: una
        # sola aggregazione per TUTTI i task in scope, poi lookup per riga nel
        # loop. Questo endpoint gira anche sulla lista di tutti i progetti
        # attivi, quindi una chiamata per task sarebbe un N+1 su ~100 task.
        # Il dict contiene solo i task calcolabili: `.get()` restituisce None
        # per i task mai scomposti o senza ore_pianificate, ed è il valore che
        # finisce nel payload (nessuna segnalazione da fare).
        # Il calcolo sta in data_db_impl come tutte le aggregazioni di dominio
        # (criticita_sforamento_progetti, margini_economia): la route non
        # ricalcola nulla, e lo stesso dato alimenta GET /api/sottotask/{id}.
        # Nota sessioni: la funzione dello strato dati apre e chiude la PROPRIA
        # sessione — per contratto, come tutte le sue sorelle. Per la durata
        # delle sue due SELECT ci sono due connessioni aperte su questo thread;
        # sono letture brevi e senza lock, e l'alternativa (passarle la
        # sessione della route) romperebbe la firma di tutto lo strato dati.
        scostamento_per_task = scostamento_stime_sottotask(task_ids_all)

        # ── 2-ter. Semaforo ritardabilità, in UNA chiamata ────────────
        # Semaforo strato 1. Stessa forma dell'innesto qui sopra: un solo
        # calcolo batch per TUTTI i progetti in scope, poi lookup per riga nel
        # loop. Sono 2 query fisse (gerarchia + sottotask) indipendenti dal
        # numero di unità — su ~110 task una chiamata per unità sarebbe un N+1.
        # La funzione cammina la gerarchia PER CONTO SUO invece di ricevere
        # quella già caricata qui: è il contratto dello strato dati (apre e
        # chiude la propria sessione, come tutte le sue sorelle), e il costo è
        # una seconda lettura degli stessi id — accettato, lo stesso che fa
        # `criticita_sforamento_progetti`. L'alternativa, passarle la sessione
        # e gli oggetti ORM di questa route, legherebbe il calcolo al suo unico
        # chiamante proprio mentre stiamo per darne un secondo
        # (GET /api/sottotask/{task_id}).
        # Dal lavoro B lo stesso prezzo si paga una seconda volta, sulla tabella
        # `sottotask`: la legge questa funzione (per i colori) e la rilegge il
        # blocco 2-quater (per le righe da serializzare). È la stessa scelta,
        # fatta di nuovo e con gli stessi occhi — non una svista: due letture di
        # una tabella piccola contro un calcolo legato al suo chiamante.
        # `oggi` non si passa: lo strato dati legge `date.today()` una volta
        # sola e lo usa per tutto l'albero.
        semaforo_per_progetto = semaforo_progetti([p.id for p in progetti])

        # ── 2-quater. Le righe dei SOTTOTASK, in UNA query ────────────
        # Lavoro B (granularità sottotask nel GANTT). Query batch su task_id +
        # lookup nel loop: lo stesso pattern di `ore_per_task`, `scostamento` e
        # `semaforo` qui sopra, e NON un `selectinload` agganciato alla catena
        # joinedload della query 1 — quella è già a tre livelli
        # (progetti → fasi → task) e appenderne un quarto la fa crescere per un
        # dato che serve solo a una minoranza dei task.
        #
        # ORDINE (ordine, id): l'id come secondo criterio rende stabile anche il
        # caso `ordine` NULL, che Postgres in ASC manda in coda. È l'ordinamento
        # di `lista_sottotask_task` in routes/sottotask.py, e le due viste dei
        # pezzi devono mostrarli nella stessa sequenza. La relationship
        # `Task.sottotask` ha già `order_by=Sottotask.ordine`, ma qui non la si
        # usa (è una query a sé), quindi l'ordine va chiesto esplicitamente.
        #
        # FILTRO «Annullato» — e la sua asimmetria con l'aggregazione, che è
        # deliberata. `semaforo_progetti` NON filtra: carica tutti i pezzi e
        # calcola un colore anche per gli annullati (che essendo fra gli stati
        # chiusi escono verdi). Qui invece si escludono, con lo stesso criterio
        # di `scostamento_stime_sottotask` — `stato != "Annullato"`, il
        # precedente della casa per «quali pezzi contano»: un pezzo annullato è
        # stato tolto dal piano e non è lavoro da disegnare in un GANTT.
        # L'asimmetria è a SENSO UNICO e per questo innocua: il payload è un
        # SOTTOINSIEME dell'albero, quindi ogni pezzo serializzato trova il suo
        # colore e nessuno esce con `semaforo: null` (il contrario — un pezzo nel
        # payload senza nodo — sarebbe il caso da temere, e non può accadere).
        # Allineare anche l'aggregazione non cambierebbe NESSUN colore: un
        # annullato è verde, e il verde non vince mai il peggio-dei-figli.
        # «Eliminato» non compare nel filtro perché per un sottotask non esiste:
        # gli stati ammessi sono STATI_PIANIFICAZIONE_SOTTOTASK («Da iniziare»,
        # «Sospeso», «Annullato»), con CHECK ck_sottotask_stato_pianificazione a
        # livello DB. Filtrarlo suggerirebbe uno stato che il modello non ha.
        # I «Sospeso» restano: sono in pausa ma ancora nel piano, come in
        # `scostamento_stime_sottotask`.
        pezzi_per_task = {}
        if task_ids_all:
            for st in (session.query(Sottotask)
                       .filter(Sottotask.task_id.in_(task_ids_all),
                               Sottotask.stato != "Annullato")
                       .order_by(Sottotask.ordine, Sottotask.id)
                       .all()):
                pezzi_per_task.setdefault(st.task_id, []).append(st)

        # ── 3. Cache nomi dipendenti per evitare lookup ripetuti ──────
        dip_rows = session.query(Dipendente).all()
        nomi_dip = {d.id: d.nome for d in dip_rows}

        # ── 4. Costruzione struttura nidificata ──────────────────────
        result = []
        for p in progetti:
            # Filtra fasi: nessun filtro qui, mostriamo tutte le fasi del progetto.
            # Il frontend decide quali aprire/chiudere per default.
            fasi_serial = []
            ore_vendute_proj = 0.0
            ore_consumate_proj = 0.0

            # Il sotto-albero del semaforo di QUESTO progetto. `or {}` così i
            # tre `.get()` più sotto restano leciti anche se il progetto non
            # comparisse: vedi `_semaforo_payload` per il perché non si alza.
            albero_semaforo = semaforo_per_progetto.get(p.id) or {}
            semaforo_fasi = albero_semaforo.get("fasi", {})

            for f in sorted(p.fasi, key=lambda x: x.ordine or 0):
                semaforo_task = semaforo_fasi.get(f.id, {}).get("task", {})
                tasks_serial = []
                ore_consumate_fase = 0.0
                # Filtra task "Eliminato" (soft delete, Step 2.4-bis fix):
                # non devono comparire nel drill-down, ma rimangono in DB.
                for t in f.task:
                    if t.stato == "Eliminato":
                        continue
                    # Il nodo del semaforo di QUESTO task, con dentro — sui soli
                    # task scomposti — il sotto-dict {sottotask_id: nodo} che il
                    # blocco più sotto appaia alle righe vere.
                    nodo_task = semaforo_task.get(t.id) or {}
                    semaforo_pezzi = nodo_task.get("sottotask", {})
                    ore_cons_t = ore_per_task.get(t.id, 0.0)
                    ore_consumate_fase += ore_cons_t
                    tasks_serial.append({
                        "id": t.id,
                        "nome": t.nome,
                        "stato": t.stato,
                        "ore_stimate": int(t.ore_stimate) if t.ore_stimate else 0,
                        # Step 2.3: il piano CORRENTE del task, finora esposto
                        # solo a livello fase. È il termine di confronto dello
                        # scostamento qui sotto: quando quello è null per
                        # assenza di piano, questo campo mostra il perché.
                        "ore_pianificate": float(t.ore_pianificate) if t.ore_pianificate is not None else None,
                        "ore_consumate": round(ore_cons_t, 1),
                        # Step 2.3: scostamento fra la somma delle stime dei
                        # sottotask e il piano del task, o null se non c'è
                        # niente da segnalare. SEGNALA, NON IMPONE: sono tre
                        # numeri, nessun giudizio — coerente col design
                        # "endpoint stupido, il frontend decide la resa".
                        # Stessa forma di GET /api/sottotask/{task_id}.
                        "scostamento": scostamento_per_task.get(t.id),
                        # Semaforo strato 1: colore + provenienza. Su un task
                        # `origine` è "propria" o None finché i sottotask non
                        # esistono; con i pezzi diventa "figli"/"entrambe", e il
                        # colore ne tiene conto — i pezzi che lo hanno prodotto
                        # sono annidati qui sotto, nella chiave `sottotask`.
                        "semaforo": _semaforo_payload(nodo_task or None),
                        "data_inizio": t.data_inizio.isoformat() if t.data_inizio else None,
                        "data_fine": t.data_fine.isoformat() if t.data_fine else None,
                        "dipendente_id": t.dipendente_id or "",
                        "dipendente_nome": nomi_dip.get(t.dipendente_id, ""),
                        "profilo_richiesto": t.profilo_richiesto or "",
                        # Step 3.1 (Gruppo A): predecessore principale dalla
                        # tabella-grafo, non più dal campo stringa rimosso.
                        # Mantenuto come fallback per i consumatori in transizione.
                        "predecessore": _predecessore_principale(t),
                        # Step 3.1 (Gruppo B): lista completa delle dipendenze
                        # entranti tipizzate (FS/SS/FF/SF). Il frontend del
                        # Cantiere la usa per mostrare la cascata della fase e
                        # popolare il form dipendenze multiple in modifica.
                        # selectinload(Task.dipendenze_entranti) è già caricato
                        # nella query sopra (no N+1).
                        "dipendenze": [
                            {"task_predecessore_id": d.task_predecessore_id,
                             "tipo_dipendenza": d.tipo_dipendenza}
                            for d in (t.dipendenze_entranti or [])
                        ],
                    })

                    # ── I SOTTOTASK (lavoro B) ────────────────────────
                    # La chiave compare SOLO sui task scomposti: la sua assenza
                    # è essa stessa l'informazione «questo task non ha pezzi»,
                    # e il frontend distingue i due render da lì. È la
                    # convenzione della casa (`task_settimana_dipendente`,
                    # `scostamento_stime_sottotask`), e tiene il payload dei
                    # task non scomposti — cioè tutti e 114 quelli in DB oggi —
                    # IDENTICO a prima: nessun consumatore esistente vede
                    # comparire un campo nuovo. Una lista vuota sempre presente
                    # direbbe la stessa cosa in modo più rumoroso, e cambierebbe
                    # il payload di ogni task del sistema per niente.
                    # `tasks_serial[-1]` invece di costruire prima la lista: è
                    # come lo fa `task_settimana_dipendente` (`out[-1][...]`),
                    # ed evita di spezzare il dict literal qui sopra.
                    pezzi = pezzi_per_task.get(t.id)
                    if pezzi:
                        tasks_serial[-1]["sottotask"] = [
                            {
                                "id": st.id,
                                "nome": st.nome,
                                "ordine": st.ordine,
                                "stato": st.stato,
                                "ore_stimate": st.ore_stimate,
                                # RAW, e `None` non "" — al contrario del task
                                # qui sopra, che usa `t.dipendente_id or ""`.
                                # Non è un'incoerenza: su un sottotask il NULL È
                                # informazione («nessun override, lo fa chi fa
                                # il task»), mentre su un task "" vuol dire solo
                                # «non assegnato». Appiattire il NULL sul nome
                                # dell'assegnatario del task nasconderebbe quali
                                # pezzi hanno un override esplicito — la stessa
                                # scelta, con la stessa motivazione, di
                                # `lista_sottotask_task` in routes/sottotask.py,
                                # di cui questo payload riusa il vocabolario
                                # invece di inventarne un secondo.
                                "dipendente_id": st.dipendente_id,
                                "dipendente_nome": nomi_dip.get(st.dipendente_id, ""),
                                # Colore GIÀ CALCOLATO da `semaforo_progetti`,
                                # non ricalcolato qui: l'aggregazione scende ai
                                # sottotask per fare il peggio-dei-figli del
                                # task, e fin qui il suo risultato veniva
                                # scartato. Il pezzo eredita la finestra
                                # temporale del padre (vedi sotto), quindi il
                                # suo colore riflette la data DEL TASK.
                                "semaforo": _semaforo_payload(semaforo_pezzi.get(st.id)),
                            }
                            for st in pezzi
                        ]

                ore_vendute_fase = float(f.ore_vendute or 0)
                ore_vendute_proj += ore_vendute_fase
                ore_consumate_proj += ore_consumate_fase

                fasi_serial.append({
                    "id": f.id,
                    "nome": f.nome,
                    "ordine": f.ordine,
                    "stato": f.stato,
                    "data_inizio": f.data_inizio.isoformat() if f.data_inizio else None,
                    "data_fine": f.data_fine.isoformat() if f.data_fine else None,
                    "ore_vendute": ore_vendute_fase,
                    "ore_pianificate": float(f.ore_pianificate or 0),
                    "ore_consumate": round(ore_consumate_fase, 1),
                    # URGENZA — GREZZA e RISOLTA, due campi distinti (A1).
                    # `urgenza` null = «eredita dal progetto», ed è il caso
                    # normale; `urgenza_risolta` è il valore che vale davvero.
                    # Si espongono ENTRAMBE perché il Cantiere deve poter
                    # mostrare «(dal progetto)» invece di un livello scelto:
                    # appiattire tutto sul risolto nasconderebbe al PM quali
                    # fasi ha deciso lui. È la stessa scelta, con la stessa
                    # motivazione, di `lista_sottotask_task` sull'assegnatario.
                    # La risoluzione NON si riscrive qui: `urgenza_fase_risolta`
                    # è la regola, e sta in models.
                    "urgenza": f.urgenza,
                    "urgenza_risolta": urgenza_fase_risolta(f, p),
                    # Semaforo di fase: il peggio fra il proprio calendario e i
                    # task. `figli_rossi` = quanti task rossi ha dentro.
                    "semaforo": _semaforo_payload(semaforo_fasi.get(f.id)),
                    "n_task": len(tasks_serial),
                    "tasks": tasks_serial,
                })

            result.append({
                "id": p.id,
                "nome": p.nome,
                "cliente": p.cliente,
                "stato": p.stato,
                "stato_derivato": p.stato_derivato,  # property calcolata
                "tipologia": p.tipologia or "ordinario",
                "data_inizio": p.data_inizio.isoformat() if p.data_inizio else None,
                "data_fine": p.data_fine.isoformat() if p.data_fine else None,
                "budget_ore": int(p.budget_ore) if p.budget_ore else 0,
                "pm_id": p.pm_id,
                "ore_vendute_totali": ore_vendute_proj,
                "ore_consumate_totali": round(ore_consumate_proj, 1),
                # Urgenza del progetto: un campo solo, sempre valorizzato — è
                # NOT NULL perché è la RADICE dell'eredità delle fasi.
                "urgenza": p.urgenza,
                # Semaforo di progetto: il peggio fra il proprio calendario e
                # le fasi. `figli_rossi` = quante FASI rosse (non quanti task:
                # i diretti, vedi `_nodo_semaforo`).
                "semaforo": _semaforo_payload(albero_semaforo or None),
                "n_fasi": len(fasi_serial),
                "fasi": fasi_serial,
            })

        return result
    finally:
        session.close()


# ─────────────────────────────────────────────────────────────────────────
# Helper privato: carica i task per l'export (PDF/PNG/Excel) da Postgres
# e restituisce una lista di dict pronti all'uso. Centralizza il filtro
# "progetti non Sospesi" e la pre-conversione delle date a datetime
# (necessaria per il foglio "GANTT Visivo" — vedi NOTE TECNICHE).
# ─────────────────────────────────────────────────────────────────────────

def _carica_task_export(
    progetto_id: Optional[str],
    *,
    escludi_eliminati: bool,
    include_fase: bool = False,
) -> list[dict]:
    """Carica i task per gli endpoint di export con tutti i campi serviti.

    Args:
        progetto_id: filtro su singolo progetto. Se None, esclude i
            progetti in stato 'Sospeso' (replica del comportamento DataFrame).
        escludi_eliminati: se True, scarta i task con stato 'Eliminato'.
            PDF/PNG storicamente NON filtravano; Excel sì — passare il
            flag corretto preserva l'iso-comportamento per ciascun endpoint.
        include_fase: se True, joinedload anche su Task.fase_rel per
            poter leggere il nome fase (usato solo dall'export Excel).
    """
    session = get_session()
    opts = [joinedload(Task.progetto)]
    if include_fase:
        opts.append(joinedload(Task.fase_rel))
    q = session.query(Task).options(*opts)
    if progetto_id:
        q = q.filter(Task.progetto_id == progetto_id)
    else:
        q = q.filter(Task.progetto.has(Progetto.stato != "Sospeso"))
    if escludi_eliminati:
        q = q.filter(Task.stato != "Eliminato")
    tasks = q.all()
    session.close()

    records = []
    for t in tasks:
        records.append({
            "id": t.id,
            "nome": t.nome,
            "stato": t.stato,
            "ore_stimate": int(t.ore_stimate or 0),
            # Pre-converto a datetime per coerenza con lo storico pandas.Timestamp:
            # i confronti e i `hasattr(x, "date")` del foglio GANTT Visivo
            # dipendono dal tipo `datetime`, non `date`.
            "data_inizio": _to_dt(t.data_inizio),
            "data_fine": _to_dt(t.data_fine),
            "dipendente_id": t.dipendente_id or "",
            "progetto_id": t.progetto_id,
            "progetto_nome": t.progetto.nome if t.progetto else "?",
            "fase": (t.fase_rel.nome if t.fase_rel else "") if include_fase else "",
        })
    return records


# ═════════════════════════════════════════════════════════════════════════
# 2. GET /api/gantt/export-pdf — esportazione PDF
# ═════════════════════════════════════════════════════════════════════════

@router.get("/export-pdf")
def export_gantt_pdf(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Genera e scarica un PDF del GANTT."""
    from gantt_pdf import genera_gantt_pdf

    # NB: PDF/PNG storicamente non filtravano "Eliminato" (asimmetria col
    # GET /api/gantt e l'Excel). Iso-comportamento → escludi_eliminati=False.
    records = _carica_task_export(progetto_id, escludi_eliminati=False)

    # PDF e PNG sono grafici su asse temporale: un task senza date non è
    # rappresentabile lì e farebbe crashare gantt_pdf.py su
    # datetime.strptime("", "%Y-%m-%d"). Lo scartiamo qui. I task senza
    # date restano comunque visibili in GET /api/gantt (con start/end "")
    # e nell'Excel (riga senza barra).
    records = [r for r in records if r["data_inizio"] and r["data_fine"]]

    gantt_data = []
    for r in records:
        dip = get_dipendente(r["dipendente_id"])
        gantt_data.append({
            "id": r["id"],
            "name": r["nome"],
            # Date opzionali: task senza data passa con stringa vuota
            # al generatore PDF (evita AttributeError sul .strftime di None).
            "start": r["data_inizio"].strftime("%Y-%m-%d") if r["data_inizio"] else "",
            "end": r["data_fine"].strftime("%Y-%m-%d") if r["data_fine"] else "",
            "project": r["progetto_nome"],
            "assignee": dip["nome"],
            "status": r["stato"],
            "estimated_hours": r["ore_stimate"],
        })

    # Titolo
    if progetto_id:
        session = get_session()
        prog = session.query(Progetto).filter(Progetto.id == progetto_id).first()
        session.close()
        titolo = f"GANTT — {prog.nome}" if prog else "GANTT"
    else:
        titolo = "GANTT IMC-Group — Tutti i progetti"

    pdf_bytes = genera_gantt_pdf(gantt_data, titolo=titolo, progetto_filtro=progetto_id)
    filename = f"gantt_{progetto_id or 'tutti'}_{datetime.now().strftime('%Y%m%d')}.pdf"

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ═════════════════════════════════════════════════════════════════════════
# 3. GET /api/gantt/export-png — esportazione PNG (via PDF + pdftoppm)
# ═════════════════════════════════════════════════════════════════════════

@router.get("/export-png")
def export_gantt_png(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Genera e scarica un PNG del GANTT (convertendo il PDF)."""
    from gantt_pdf import genera_gantt_pdf
    import subprocess
    import tempfile
    import os

    # Stessa logica di export-pdf (storicamente non filtrava "Eliminato").
    records = _carica_task_export(progetto_id, escludi_eliminati=False)

    # Stesso filtro di export-pdf: task senza date non sono rappresentabili
    # su asse temporale (vedi nota in export_gantt_pdf).
    records = [r for r in records if r["data_inizio"] and r["data_fine"]]

    gantt_data = []
    for r in records:
        dip = get_dipendente(r["dipendente_id"])
        gantt_data.append({
            "id": r["id"], "name": r["nome"],
            # Date opzionali: stesso pattern di export-pdf.
            "start": r["data_inizio"].strftime("%Y-%m-%d") if r["data_inizio"] else "",
            "end": r["data_fine"].strftime("%Y-%m-%d") if r["data_fine"] else "",
            "project": r["progetto_nome"], "assignee": dip["nome"],
            "status": r["stato"], "estimated_hours": r["ore_stimate"],
        })

    if progetto_id:
        session = get_session()
        prog = session.query(Progetto).filter(Progetto.id == progetto_id).first()
        session.close()
        titolo = f"GANTT — {prog.nome}" if prog else "GANTT"
    else:
        titolo = "GANTT IMC-Group — Tutti i progetti"

    pdf_bytes = genera_gantt_pdf(gantt_data, titolo=titolo)

    # Converti PDF → PNG con pdftoppm (poppler-utils)
    try:
        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp_pdf:
            tmp_pdf.write(pdf_bytes)
            tmp_pdf_path = tmp_pdf.name

        png_path = tmp_pdf_path.replace(".pdf", "")
        subprocess.run(
            ["pdftoppm", "-png", "-r", "200", "-singlefile", tmp_pdf_path, png_path],
            check=True, capture_output=True
        )

        png_file = png_path + ".png"
        with open(png_file, "rb") as f:
            png_bytes = f.read()

        os.unlink(tmp_pdf_path)
        os.unlink(png_file)

        filename = f"gantt_{progetto_id or 'tutti'}_{datetime.now().strftime('%Y%m%d')}.png"
        return Response(
            content=png_bytes,
            media_type="image/png",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except (subprocess.CalledProcessError, FileNotFoundError):
        raise HTTPException(
            500,
            "Conversione PNG non disponibile. Installa poppler-utils: sudo apt install poppler-utils"
        )


# ═════════════════════════════════════════════════════════════════════════
# 4. GET /api/gantt/export-excel — esportazione Excel (2 fogli)
# ═════════════════════════════════════════════════════════════════════════

@router.get("/export-excel")
def export_gantt_excel(
    progetto_id: Optional[str] = None,
    _: Utente = Depends(require_manager),
):
    """Genera e scarica un file Excel con i dati GANTT + foglio visivo."""
    import io
    import pandas as pd

    # Excel storicamente filtrava "Eliminato".
    records = _carica_task_export(progetto_id, escludi_eliminati=True, include_fase=True)

    # Costruisci dati per il foglio "Dati GANTT"
    export_data = []
    for r in records:
        dip = get_dipendente(r["dipendente_id"])
        export_data.append({
            "Progetto": r["progetto_nome"],
            "Task": r["nome"],
            "Fase": r["fase"],
            "Assegnato a": dip["nome"],
            "Profilo": dip["profilo"],
            "Ore stimate": r["ore_stimate"],
            # Datetime: hasattr(.,"strftime") sempre True. Equivalente a passare il valore.
            "Data inizio": r["data_inizio"],
            "Data fine": r["data_fine"],
            "Stato": r["stato"],
        })

    df = pd.DataFrame(export_data)
    df = df.sort_values(["Progetto", "Data inizio"])

    buffer = io.BytesIO()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        wb = Workbook()

        # ═══ FOGLIO 1: Tabella dati ═══
        ws_data = wb.active
        ws_data.title = "Dati GANTT"

        headers = ["Progetto", "Task", "Fase", "Assegnato a", "Profilo",
                   "Ore stimate", "Data inizio", "Data fine", "Stato"]
        header_fill = PatternFill(start_color="1a365d", end_color="1a365d", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)

        for col, h in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Dati con colore per stato
        for row_idx, (_, row) in enumerate(df.iterrows(), 2):
            for col_idx, h in enumerate(headers, 1):
                val = row[h]
                if h in ("Data inizio", "Data fine"):
                    # pd.notna() copre sia None che NaT (a differenza di
                    # hasattr(., "strftime") che è True su NaT e fa crashare
                    # NaT.strftime con ValueError). Date assenti → cella vuota.
                    val = val.strftime("%d/%m/%Y") if pd.notna(val) else None
                cell = ws_data.cell(row=row_idx, column=col_idx, value=val)
                cell.alignment = Alignment(horizontal="left")

                if row["Stato"] == "Completato":
                    cell.fill = PatternFill(start_color="d4edda", end_color="d4edda", fill_type="solid")
                elif row["Stato"] == "Da iniziare":
                    cell.fill = PatternFill(start_color="e2e8f0", end_color="e2e8f0", fill_type="solid")

        # Auto-width
        for col in range(1, len(headers) + 1):
            max_len = max(
                len(str(ws_data.cell(row=r, column=col).value or ""))
                for r in range(1, ws_data.max_row + 1)
            )
            ws_data.column_dimensions[get_column_letter(col)].width = min(35, max(10, max_len + 2))

        # ═══ FOGLIO 2: GANTT Visivo ═══
        ws_gantt = wb.create_sheet("GANTT Visivo")

        # Trova range date — records ha già le date come datetime (_to_dt)
        date_inizio = [r["data_inizio"] for r in records if r["data_inizio"] is not None]
        date_fine = [r["data_fine"] for r in records if r["data_fine"] is not None]

        if not date_inizio:
            ws_gantt.cell(row=1, column=1, value="Nessun task da visualizzare")
        else:
            from datetime import timedelta as td
            min_date = min(date_inizio)
            max_date = max(date_fine)

            # Genera settimane (lunedì del primo task → venerdì dell'ultimo + 1 sett)
            settimane = []
            current = min_date - td(days=min_date.weekday())  # lunedì
            while current <= max_date + td(days=7):
                settimane.append(current)
                current += td(days=7)

            n_sett = len(settimane)

            # Colori per progetto (mantenuti per coerenza con vista frontend)
            colori_progetto = {
                "Adeguamento DORA": "4472C4",
                "Framework Compliance 262": "ED7D31",
                "Digitalizzazione Corpo Normativo": "70AD47",
                "Risk Assessment Operativo": "FFC000",
                "ProcessBook Aziendale": "9B59B6",
                "Attività Interne": "95A5A6",
            }
            colore_default = "5B9BD5"

            # Colori per stato
            colori_stato = {
                "Completato": "27ae60",
                "In corso": "3498db",
                "Da iniziare": "95a5a6",
                "Sospeso": "e67e22",
            }

            # Header
            ws_gantt.cell(row=1, column=1, value="Task").font = Font(bold=True, size=9)
            ws_gantt.cell(row=1, column=2, value="Risorsa").font = Font(bold=True, size=9)
            ws_gantt.cell(row=1, column=3, value="Progetto").font = Font(bold=True, size=9)
            ws_gantt.cell(row=1, column=4, value="Stato").font = Font(bold=True, size=9)

            for i, sett in enumerate(settimane):
                cell = ws_gantt.cell(row=1, column=5 + i, value=sett.strftime("%d/%m"))
                cell.font = Font(size=7, bold=True)
                cell.alignment = Alignment(horizontal="center", text_rotation=90)
                ws_gantt.column_dimensions[get_column_letter(5 + i)].width = 4

            ws_gantt.column_dimensions["A"].width = 30
            ws_gantt.column_dimensions["B"].width = 18
            ws_gantt.column_dimensions["C"].width = 22
            ws_gantt.column_dimensions["D"].width = 12

            # Header fill
            for col in range(1, 5 + n_sett):
                ws_gantt.cell(row=1, column=col).fill = PatternFill(
                    start_color="2c3e50", end_color="2c3e50", fill_type="solid"
                )
                ws_gantt.cell(row=1, column=col).font = Font(color="FFFFFF", bold=True, size=8)

            # Righe task — sort Python-side per (progetto_id, data_inizio).
            # Task senza data_inizio → in fondo al loro progetto (datetime.max),
            # altrimenti sorted() solleva TypeError confrontando None vs datetime.
            records_ordinati = sorted(
                records,
                key=lambda r: (r["progetto_id"], r["data_inizio"] or datetime.max)
            )
            row = 2
            current_project = ""
            for r in records_ordinati:
                dip = get_dipendente(r["dipendente_id"])
                proj_nome = r["progetto_nome"]

                # Riga separatore progetto
                if proj_nome != current_project:
                    current_project = proj_nome
                    sep_cell = ws_gantt.cell(row=row, column=1, value=proj_nome.upper())
                    sep_cell.font = Font(bold=True, size=9, color="FFFFFF")
                    proj_color = colori_progetto.get(proj_nome, colore_default)
                    for col in range(1, 5 + n_sett):
                        ws_gantt.cell(row=row, column=col).fill = PatternFill(
                            start_color=proj_color, end_color=proj_color, fill_type="solid"
                        )
                    row += 1

                # Task info
                ws_gantt.cell(row=row, column=1, value=r["nome"]).font = Font(size=9)
                ws_gantt.cell(row=row, column=2, value=dip["nome"]).font = Font(size=8, color="666666")
                ws_gantt.cell(row=row, column=3, value=proj_nome).font = Font(size=8, color="666666")
                ws_gantt.cell(row=row, column=4, value=r["stato"]).font = Font(size=8)

                # Colore stato nella cella stato
                stato_color = colori_stato.get(r["stato"], "95a5a6")
                ws_gantt.cell(row=row, column=4).fill = PatternFill(
                    start_color=stato_color, end_color=stato_color, fill_type="solid"
                )
                ws_gantt.cell(row=row, column=4).font = Font(size=8, color="FFFFFF")

                # Barre GANTT — confronto datetime vs datetime (date pre-convertite)
                task_start = r["data_inizio"]
                task_end = r["data_fine"]
                bar_color = colori_stato.get(r["stato"], colore_default)

                # Se manca anche una sola delle due date, nessuna barra:
                # la riga task compare comunque (info + stato colorato),
                # niente confronto None vs datetime che solleverebbe TypeError.
                if task_start is not None and task_end is not None:
                    for i, sett in enumerate(settimane):
                        sett_end = sett + td(days=6)
                        if task_start <= sett_end and task_end >= sett:
                            ws_gantt.cell(row=row, column=5 + i).fill = PatternFill(
                                start_color=bar_color, end_color=bar_color, fill_type="solid"
                            )

                row += 1

            # Legenda in fondo
            row += 2
            ws_gantt.cell(row=row, column=1, value="Legenda:").font = Font(bold=True, size=9)
            row += 1
            for stato, colore in colori_stato.items():
                ws_gantt.cell(row=row, column=1, value=stato).font = Font(size=9)
                ws_gantt.cell(row=row, column=2).fill = PatternFill(
                    start_color=colore, end_color=colore, fill_type="solid"
                )
                row += 1

        wb.save(buffer)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ext = "xlsx"

    except ImportError:
        # Fallback CSV
        buffer = io.BytesIO()
        df_export = df.copy()
        # pd.notna() copre None e NaT (vedi nota nel foglio "Dati GANTT").
        # Date assenti → stringa vuota nel CSV.
        df_export["Data inizio"] = df_export["Data inizio"].apply(
            lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) else ""
        )
        df_export["Data fine"] = df_export["Data fine"].apply(
            lambda x: x.strftime("%d/%m/%Y") if pd.notna(x) else ""
        )
        df_export.to_csv(buffer, index=False, sep=";", encoding="utf-8-sig")
        media_type = "text/csv"
        ext = "csv"

    buffer.seek(0)
    filename = f"gantt_{progetto_id or 'tutti'}_{datetime.now().strftime('%Y%m%d')}.{ext}"
    return Response(
        content=buffer.getvalue(),
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
